"""
PO Cost Changes — pandas pipeline for the new TicketVault DB export.

This is the v2 source: a single-header sheet (one row per ticket-group line)
with richer fields than the old PO_Cost_Changes.xlsm export. The reader
(`_read_one`) normalizes each new file into the exact internal schema the
proven pipeline already expects, so the transform / Bills / Expenses / Summary
logic below is unchanged from the original Section1.m port (see ../docs).

Differences from v1, by design:
  * No Purchase Details uploads. The new export carries its own
    PurchaseDetailMatchFound flag, but per spec we IGNORE it and process all
    rows. Row exclusion is driven solely by the manual "Remove" = X flag.
  * Total Adjustment is DERIVED = TicketCostTotal - InitialTicketCostTotal
    (End - Start); the source no longer ships an explicit adjustment column.
  * AdjustedDateTimeUTC is in UTC; we convert to US Central before taking the
    date used for bucketing and the date-range badge.
"""
from __future__ import annotations

import re
import random
from datetime import date as _date

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from mapping import get_mapping
from teams import get_teams
from vendors import get_tc_vendors, offset_category
from vendor_rules import apply_vendor_pipeline, clean_ext_po

# QBO Company → Company-column value (what appears in the Company column of the
# output data / QBO import).
DISPLAY_NAMES: dict[str, str] = {
    "Y&S Tickets":         "Y&S",
    "Damona & Crew":       "Damona",
    "The Ticket Guy LLC":  "Ticket Guy",
    "YourTickets":         "YourTickets",
    "YS Asher Tickets":    "YSA",
    "YS Chase Tickets":    "Chase (Jacks)",
    "YS Katz Tickets":     "Katz",
    "YS Levine Tickets":   "Levine",
    "YS Levovitz Tickets": "Levovitz",
    "YS Needle Tickets":   "Needle",
    "YS TL Tickets":       "TL",
    "YSKG Tickets":        "YSKG",
    "YSM Tickets":         "Grossman",
    "YSP Tickets":         "Pollak",
    "YSS Tickets":         "Sternbuch",
    "YSW Tickets":         "YSW (Waxler)",
}

# Company-column value → short label used for the per-company sheet/tab names,
# download file names, and the UI grid. Mirrors the Purchase Details app, where
# the sheet label and the in-data Company value differ for these companies
# (e.g. sheet "GK" but Company value "YSKG"). Anything not listed uses its
# Company value as the label.
FILE_LABELS: dict[str, str] = {
    "YSKG":          "GK",
    "Chase (Jacks)": "Chase",
    "YSW (Waxler)":  "Waxler",
    # "Ticket Guy" — sheet label and Company value are the same.
}


def file_label(company_value: str) -> str:
    """Short sheet/file/UI label for a Company-column value."""
    return FILE_LABELS.get(company_value, company_value)


# Raw TicketVault company → the value shown in the Company column. The Purchase
# Details app renames the Company value in-data for just these companies;
# everyone else keeps their raw company name (e.g. "YS-Seatgeek2"). Keys are
# lowercased for case-insensitive matching.
COMPANY_VALUE_RENAMES: dict[str, str] = {
    "gk llc":                 "YSKG",
    "the ticket guy":         "Ticket Guy",
    "the ticket guy-jas":     "Ticket Guy",
    "the ticket guy-legacy":  "Ticket Guy",
    "the ticket guy vip":     "Ticket Guy",
    "jacks ys":               "Chase (Jacks)",
    "ysw":                    "YSW (Waxler)",
}


def company_value(original_company) -> str:
    """The Company-column value for a raw TicketVault company name. Renames the
    PD-renamed companies; passes everything else through unchanged."""
    if original_company is None or (isinstance(original_company, float) and pd.isna(original_company)):
        return original_company
    return COMPANY_VALUE_RENAMES.get(str(original_company).strip().lower(), original_company)

# Sheet/file order for the combined workbook and UI grid. Matches the
# Purchase Details processor's ordering. Any display label not listed here
# falls to the end (alphabetical), and "YourTickets" always goes last.
DISPLAY_ORDER: list[str] = [
    "Y&S",
    "Grossman",
    "Sternbuch",
    "Pollak",
    "Levine",
    "Levovitz",
    "GK",
    "Ticket Guy",
    "Chase",
    "YSA",
    "Katz",
    "Needle",      # not in reference; slotted with the other affiliates
    "TL",
    "Waxler",
    "Damona",
    "YourTickets",  # always last
]


def display_name(qbo_company: str) -> str:
    """Return the display label for a QBO company, or the QBO name itself
    if there's no override. Lets new QBO companies in the master file work
    without a code change — they just don't get a custom label."""
    return DISPLAY_NAMES.get(qbo_company, qbo_company)


def _sort_key(label: str) -> tuple[int, str]:
    """Sort key that respects DISPLAY_ORDER, with unknown labels appended
    alphabetically just before YourTickets (which always goes last).

    Returns a (priority, label) tuple:
      - YourTickets → (2, '')               always last
      - Listed labels → (0, position)       in DISPLAY_ORDER index order
      - Unknown labels → (1, label)         alphabetically before YourTickets
    """
    if label == "YourTickets":
        return (2, "")
    try:
        return (0, f"{DISPLAY_ORDER.index(label):03d}")
    except ValueError:
        return (1, label)

# Columns dropped early (seat-level detail). Listed once so the schema is obvious.
SEAT_LEVEL_COLUMNS = [
    "Opponent/Performer", "Event Date", "Seat Section", "Seat Row", "Seats",
    "Ticket Cost Start", "Ticket Cost End", "Qty Start", "Qty End",
]

# Final output column order. "Original Company" is a helper used by the
# Bills/Expenses builders. Other helpers (e.g. "_display_label") are
# stripped before writing visible sheets.
FINAL_COLUMNS = [
    "Company", "Adjustment Date", "Vendor", "Team/Performer",
    "AccountEmail", "ExtPONumber", "Total Adjustment", "Original Company",
    "CreatedDate",
]

# Aggregation key for the final collapse.
# Rows that share (Company, Adjustment Date, Vendor, Team/Performer,
# AccountEmail, ExtPONumber) get summed; otherwise they stay separate.
AGGREGATION_KEYS = [
    "Company", "Adjustment Date", "Vendor", "Team/Performer",
    "AccountEmail", "ExtPONumber",
]

# Group keys for the FIRST collapse — one row per PO event — used by the
# cancellation override logic, which needs Cancelled and Total End.
# AccountEmail/ExtPONumber are carried here so they survive to the final
# aggregation (and into the Bills/Expenses memo).
GROUP_KEYS = [
    "Company", "Original Company", "PO #", "Adjustment Date", "Vendor",
    "Team/Performer", "AccountEmail", "ExtPONumber", "Cancelled", "User",
]


def _collect_created_dates(s) -> tuple:
    """Stage-1 aggregation: the distinct non-null created dates (date-only) in a
    PO-event group, sorted ascending."""
    vals = {pd.Timestamp(x).normalize() for x in s.dropna()}
    return tuple(sorted(vals))


def _union_created_dates(s) -> tuple:
    """Stage-2 aggregation: union of the per-event created-date tuples."""
    acc: set = set()
    for t in s:
        if isinstance(t, tuple):
            acc.update(t)
        elif t is not None and not (isinstance(t, float) and pd.isna(t)):
            acc.add(pd.Timestamp(t).normalize())
    return tuple(sorted(acc))


def transform(
    df: pd.DataFrame,
    mapping: dict[str, str] | None = None,
) -> tuple[pd.DataFrame, dict]:
    """Run the full PO Cost Changes pipeline.

    Mirrors Power Query Section1.m → Table1. Step numbers below correspond
    to the M code's named steps.

    Args:
        df: raw input DataFrame matching the template's Table1 schema.
        mapping: optional override for the TicketVault→QBO company mapping.
                 If None, loads from Master_Mapping_List.xlsx via app.mapping.

    Returns:
        (cleaned_df, dropped_info) where dropped_info is:
            {"unmapped_companies": {"<name>": <row_count>, ...},
             "total_dropped_rows": int}
        The dropped info lets the UI flag accidental uploads of non-QBO companies.
    """
    if mapping is None:
        mapping = get_mapping()

    # Guard: if every row was excluded upstream, there's nothing to transform.
    # Return an empty frame with the final columns so downstream sheet/stat
    # builders work uniformly.
    if df.empty:
        empty = pd.DataFrame(columns=FINAL_COLUMNS)
        return empty, {"unmapped_companies": {}, "total_dropped_rows": 0}

    # 1. Source — caller already loaded the data.
    out = df.copy()

    # 2. Changed Type — coerce the columns that matter for math/comparison.
    out["PO #"] = pd.to_numeric(out["PO #"], errors="coerce").astype("Int64")
    out["Adjustment Date"] = pd.to_datetime(out["Adjustment Date"], errors="coerce")
    # Strip the time component — the source data has timestamps but every
    # row we care about is a per-day event, and the aggregation key
    # depends on dates matching exactly.
    out["Adjustment Date"] = out["Adjustment Date"].dt.normalize()
    for col in ["Ticket Cost Total Start", "Ticket Cost Total End",
                "Per Ticket Adjustment", "Total Adjustment"]:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce")
    for col in ["Company", "Vendor", "Team/Performer", "Cancelled", "User"]:
        if col in out.columns:
            out[col] = out[col].astype("string")
    # Cancelled feeds into boolean comparisons later; treat NaN/NA as empty string.
    if "Cancelled" in out.columns:
        out["Cancelled"] = out["Cancelled"].fillna("")

    # 3. Removed Columns — drop seat-level detail.
    out = out.drop(columns=[c for c in SEAT_LEVEL_COLUMNS if c in out.columns])

    # 4. Renamed Columns
    out = out.rename(columns={
        "Ticket Cost Total Start": "Total Start",
        "Ticket Cost Total End": "Total End",
    })

    # 4b. Vendor renaming is now handled by the full multi-stage pipeline
    #     (vendor_rules.apply_vendor_pipeline), applied at step 6e below —
    #     after company mapping, so Original Company is available for the
    #     company-gated rules (YSA, Ticket Guy), and before the first groupby,
    #     so renamed vendors participate in aggregation.

    # 5. Changed Type1 (Currency) — already numeric; skip.
    # 5b. Removed Columns1 — drop per-ticket adjustment.
    out = out.drop(columns=[c for c in ["Per Ticket Adjustment"] if c in out.columns])

    # 5c. Snapshot the original Company string (trimmed, casing preserved)
    #     before any mapping. The Bills/Expenses tabs need this for QBO entry.
    out["Original Company"] = out["Company"].map(
        lambda v: str(v).strip() if pd.notna(v) else v
    )

    # 6. Replaced Value — map TicketVault company names to QBO names.
    #    Case-insensitive: master keys are lowercased at load time; we
    #    lowercase the input before lookup. Output keeps the canonical
    #    QBO casing from the master file.
    out["Company"] = out["Company"].map(
        lambda v: mapping.get(str(v).strip().lower(), v) if pd.notna(v) else v
    )

    # 6b. Filter to QBO-mapped companies only.
    #     The master file's QBO Company column is the gating list — anything
    #     else is either an upload mistake or out of scope. We record what
    #     got dropped so the response surfaces it.
    #     Membership check is case-insensitive against the canonical QBO names,
    #     so an upload that uses the canonical name in any casing is preserved.
    canonical_qbo_names = set(mapping.values())
    canonical_qbo_lower = {n.lower() for n in canonical_qbo_names}
    is_mapped = out["Company"].map(
        lambda v: pd.notna(v) and str(v).strip().lower() in canonical_qbo_lower
    )
    dropped = out[~is_mapped]
    unmapped_counts: dict[str, int] = {}
    if not dropped.empty:
        for name, grp in dropped.groupby("Company", dropna=False):
            label = "(blank)" if pd.isna(name) else str(name)
            unmapped_counts[label] = int(len(grp))
    out = out[is_mapped].reset_index(drop=True)

    # 6c. Normalize Company to the canonical casing from the master, so that
    #     two inputs that differ only in casing (e.g. "Y&S Tickets" and
    #     "y&s tickets") collapse into a single output bucket.
    canonical_by_lower = {n.lower(): n for n in canonical_qbo_names}
    out["Company"] = out["Company"].map(
        lambda v: canonical_by_lower.get(str(v).strip().lower(), v) if pd.notna(v) else v
    )

    # 6d. Swap canonical QBO name for its display label (e.g. "YSKG Tickets" -> "YSKG").
    #     The Company column in all output files now uses the short label.
    out["Company"] = out["Company"].map(display_name)

    # 6e. Order-number cleaning — runs BEFORE vendor renaming (matching the
    #     Purchase Details app, where clean_ext_po(df_raw) precedes
    #     build_all_query). So the Concert Seasons / Ticketmaster AM blanking
    #     checks the RAW vendor names, before any rename resolves them to a
    #     team or venue. Also blanks UUID / 19+ digit order numbers.
    out = clean_ext_po(out)

    # 6f. Vendor renaming — full multi-stage pipeline ported from the Purchase
    #     Details app. Keys off Original Company (raw company) and VenueName.
    #     Run per-row here, before any grouping, so renamed vendors drive both
    #     the aggregation key and the memo.
    if "VenueName" not in out.columns:
        out["VenueName"] = ""
    out = apply_vendor_pipeline(out)

    # VenueName has done its job (vendor lookups); it's not a final key.
    out = out.drop(columns=["VenueName"], errors="ignore")

    # 6b. Live Nation (concert-season) collapse. Concert Seasons / Live Nation
    #     Flex / YSA Live Nation all become a "Live Nation …" vendor. Bought as
    #     a season, so the per-event performer / email / order isn't meaningful
    #     for QBO: blank those three keys and label the detail "Various /
    #     Various" so every Live Nation row for a company (per date + vendor)
    #     aggregates into one line whose memo reads
    #     "Various / Various / Cost Changes (Company) (PO created date …)".
    #     Mirrors the Purchase Details app (Live Nation Extras included).
    if not out.empty:
        ln_mask = out["Vendor"].astype("string").str.contains("Live Nation", case=False, na=False)
        if ln_mask.any():
            out.loc[ln_mask, "Team/Performer"] = "Various / Various"
            out.loc[ln_mask, "AccountEmail"] = ""
            out.loc[ln_mask, "ExtPONumber"] = ""

    # 7. Grouped Rows — collapse seat lines into one PO event row. CreatedDate
    #    is carried as the latest (max) value in the group so it can feed the
    #    memo's "(PO created date …)" suffix.
    _agg7 = {"Total Start": "sum", "Total End": "sum", "Total Adjustment": "sum"}
    if "CreatedDate" in out.columns:
        _agg7["CreatedDate"] = _collect_created_dates
    out = (
        out.groupby(GROUP_KEYS, dropna=False, as_index=False)
        .agg(_agg7)
    )

    # 8. Added Conditional Column — for cancellations, override adjustment
    #    to the full negative of Total End (reverses the booking).
    out["Total Adjustment"] = out.apply(
        lambda r: -r["Total End"] if r["Cancelled"] == "Yes" else r["Total Adjustment"],
        axis=1,
    )

    # 9. Filtered Rows — drop zero-impact rows (per the M code's behavior;
    #    further zero-sum aggregates are filtered again after step 12).
    out = out[out["Total Adjustment"] != 0].reset_index(drop=True)

    # 11. Cancelled is no longer needed beyond this point (step 8 has already
    #     applied its side-effect). It would otherwise carry into the
    #     aggregation key and split groups artificially.
    out = out.drop(columns=["Cancelled"])

    # 12. Final aggregation — collapse all rows with matching
    #     (Company, Adjustment Date, Vendor, Team/Performer, AccountEmail,
    #     ExtPONumber) into one row, summing Total Adjustment. Drops PO #,
    #     Total Start, Total End, User. For Original Company, keep the most
    #     common value within each group (in case the same display-label bucket
    #     has rows that came in under slightly different original spellings).
    if not out.empty:
        _agg12 = {
            "Total Adjustment": "sum",
            "Original Company": lambda s: s.mode().iat[0] if not s.mode().empty else s.iloc[0],
        }
        if "CreatedDate" in out.columns:
            _agg12["CreatedDate"] = _union_created_dates
        out = (
            out.groupby(AGGREGATION_KEYS, dropna=False, as_index=False)
            .agg(_agg12)
        )

        # Filter zero-sum aggregates: separate +/− entries on the same key
        # can cancel out. They have no QBO impact, so drop them.
        out = out[out["Total Adjustment"] != 0].reset_index(drop=True)

    # 13. Sort for stable output: by display order (Company), then date.
    if not out.empty:
        out["_display_order"] = out["Company"].map(_sort_key)
        out = (
            out.sort_values(["_display_order", "Adjustment Date", "Vendor", "Team/Performer"],
                            kind="mergesort")
            .drop(columns=["_display_order"])
            .reset_index(drop=True)
        )

    cleaned = out[FINAL_COLUMNS]
    dropped_info = {
        "unmapped_companies": unmapped_counts,
        "total_dropped_rows": sum(unmapped_counts.values()),
    }
    return cleaned, dropped_info


def summarize_by_company(transformed: pd.DataFrame) -> dict:
    """Per-company row count and total cost change, plus grand totals.

    Returns:
        {
          "companies": [{"company": str, "rows": int, "total_adjustment": float}, ...],
          "totals": {"rows": int, "total_adjustment": float}
        }
    """
    if transformed.empty:
        return {"companies": [], "totals": {"rows": 0, "total_adjustment": 0.0}}

    grouped = (
        transformed.groupby("Company", dropna=False, as_index=False)
        .agg(rows=("Total Adjustment", "size"), total_adjustment=("Total Adjustment", "sum"))
    )
    # Sort by the same display order used everywhere else (YourTickets last).
    # NaN/blank Company values get pushed to the very end.
    grouped["_sort"] = grouped["Company"].map(
        lambda c: _sort_key(str(c)) if pd.notna(c) else (3, "")
    )
    grouped = grouped.sort_values("_sort").drop(columns=["_sort"])
    companies = [
        {
            "company": ("(blank)" if pd.isna(r["Company"]) else str(r["Company"])),
            "rows": int(r["rows"]),
            "total_adjustment": float(r["total_adjustment"]),
        }
        for _, r in grouped.iterrows()
    ]
    return {
        "companies": companies,
        "totals": {
            "rows": int(len(transformed)),
            "total_adjustment": float(transformed["Total Adjustment"].sum()),
        },
    }


# ---------------------------------------------------------------------------
# High-level orchestration for the web layer
# ---------------------------------------------------------------------------

import io
from datetime import date
from pathlib import Path
import logging

log = logging.getLogger(__name__)


def _ordinal(n: int) -> str:
    """1 -> '1st', 11 -> '11th', 22 -> '22nd'."""
    if 11 <= (n % 100) <= 13:
        return f"{n}th"
    return f"{n}{ {1:'st', 2:'nd', 3:'rd'}.get(n % 10, 'th') }"


def _format_date_range(dates: list[pd.Timestamp]) -> str:
    """e.g. 'May 1st thru May 3rd 2026' — same shape as the reference project."""
    valid = [d for d in dates if pd.notna(d)]
    if not valid:
        return date.today().strftime("%B %Y")
    valid = sorted(valid)

    def fmt(d):
        return f"{d.strftime('%B')} {_ordinal(d.day)}"

    if len(valid) == 1 or valid[0] == valid[-1]:
        return f"{fmt(valid[0])} {valid[0].strftime('%Y')}"
    return f"{fmt(valid[0])} thru {fmt(valid[-1])} {valid[-1].strftime('%Y')}"


# ── New-export schema ──────────────────────────────────────────────────────
# Column mapping: new TicketVault DB export → internal pipeline names.
NEW_COLUMN_MAP = {
    "CompanyName":            "Company",
    "PurchaseOrderID":        "PO #",
    "Vendor":                 "Vendor",
    "PerformerName":          "Team/Performer",
    "InitialTicketCostTotal": "Ticket Cost Total Start",
    "TicketCostTotal":        "Ticket Cost Total End",
    "UpdateUser":             "User",
}
# Columns that must be present for a file to be recognized as the new export.
NEW_REQUIRED = {
    "CompanyName", "PurchaseOrderID", "Vendor", "PerformerName",
    "InitialTicketCostTotal", "TicketCostTotal", "AdjustedDateTimeUTC",
    "IsCancelled", "UpdateUser",
}

# ── Zone 1 "Modified" reformatting ──────────────────────────────────────────
# A human-readable reformat of the raw seat-level export: rename + reorder, and
# split the cost/qty fields into Start (initial) and End (current). The five
# internal ID/match columns are dropped. Values and seat-level rows are kept
# exactly — no aggregation, no date conversion. Layout matches the user's
# Raw → Modified template (including the "Team/Perfomer" header spelling).
RAW_TO_MODIFIED = [
    ("CompanyName",            "Company"),
    ("PurchaseOrderID",        "PO #"),
    ("AdjustedDateTimeUTC",    "Adjustment Date"),
    ("Vendor",                 "Vendor"),
    ("PerformerName",          "Team/Perfomer"),
    ("SecondaryPerformerName", "Opponent/Performer"),
    ("EventDate",              "Event Date"),
    ("VenueName",              "Venue"),
    ("Section",                "Sec"),
    ("Row",                    "Row"),
    ("StartSeat",              "Start Seat"),
    ("EndSeat",                "End Seat"),
    ("ExtPONumber",            "Ext PO #"),
    ("AccountEmail",           "Account Email"),
    ("IsCancelled",            "Cancelled"),
    ("UpdateUser",             "User"),
    ("TicketCost",             "Per Ticket End"),
    ("InitialTicketCost",      "Per Ticket Start"),
    ("TicketCostTotal",        "Total Ticket End"),
    ("InitialTicketCostTotal", "Total Ticket Start"),
    ("Quantity",               "Qty End"),
    ("InitialQuantity",        "Qty Start"),
    ("CreatedDate",            "CreatedDate"),
]
MODIFIED_COLUMNS = [m for _, m in RAW_TO_MODIFIED]
# Reverse map for reading a Modified file back in Zone 2. Accept both the
# template's "Team/Perfomer" spelling and the corrected "Team/Performer".
MODIFIED_TO_RAW = {m: r for r, m in RAW_TO_MODIFIED}
MODIFIED_TO_RAW["Team/Performer"] = "PerformerName"
# Signature columns unique to the Modified layout.
_MODIFIED_SIGNATURE = {"Total Ticket Start", "Total Ticket End", "PO #"}

# Accounting-team usernames whose PO cost-change lines are typically already
# entered in QBO. The Zone 1 "Converted" sheet highlights the User column for
# these so reviewers can flag them (Remove = X) before Zone 2 processing.
#
# IMPORTANT: keep this in sync with the list shown to users in
# templates/index.html (#upload-screen-warning). Stored lowercased for
# case-insensitive matching. Note that "macohen" is intentional and distinct
# from "mcohen" (which is NOT an accounting user — see the "(not mcohen)" note
# in the UI), so matching is exact rather than substring.
ACCOUNTING_USERS: frozenset[str] = frozenset({
    "jhantz", "jhantz-2", "jhantz-3", "bblumenthal", "yklein", "awealcatch",
    "cschlesinger", "dbowden", "macohen", "gbollag", "mtawil", "sgreenhouse",
    "msadriu", "lrafuna", "fajeti", "mmeta", "lbeqa", "fhoxha", "hmorina",
})

# Highlight applied to the User column for accounting-team users (Zone 1).
ACCOUNTING_FILL = PatternFill("solid", start_color="FFEB3B")  # amber-yellow


def _is_accounting_user(value) -> bool:
    """True when `value` is a known accounting-team username (case-insensitive,
    exact match — so 'mcohen' does not match 'macohen')."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return False
    return str(value).strip().lower() in ACCOUNTING_USERS


def _is_modified_format(df: pd.DataFrame) -> bool:
    return _MODIFIED_SIGNATURE.issubset(set(df.columns))

# Timezone the UTC timestamps are converted to before the date is taken.
# Only used as a FALLBACK when the filename has no parseable date (see
# _filename_adjustment_date / the Adjustment Date override below).
LOCAL_TZ = "America/Chicago"  # US Central

# Adjustment Date is overridden to the date in the filename. That YYYY-MM-DD is
# the business date the cost changes belong to. Falls back to the row's UTC
# timestamp (converted to local time) only when the filename has no date.
DATE_FROM_FILENAME = re.compile(r"(\d{4})-(\d{2})-(\d{2})")

# Hidden helper column flagging rows where AdjustedDate == CreatedDate (same
# Central calendar day). Set in normalization, consumed by the exclusion logic
# in process_files, and dropped before any sheet is written.
EXCLUDE_SAME_DATE_COL = "_exclude_same_date"


def _filename_adjustment_date(filename: str) -> pd.Timestamp | None:
    """Return the date in the filename as a midnight Timestamp, or None if the
    filename contains no YYYY-MM-DD date."""
    m = DATE_FROM_FILENAME.search(filename or "")
    if not m:
        return None
    d = _date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return pd.Timestamp(d)

# Extra source columns carried through unchanged onto Source Data (not used by
# the aggregation pipeline). AccountEmail / ExtPONumber are handled separately
# (they're aggregation keys); CreatedDate is handled separately too (converted
# to a US-Central date-only value for display).
PASSTHROUGH_COLUMNS = ["VenueName"]


def _clean_key_str(v) -> str:
    """Normalize a key/memo value to a clean string: NaN → '', strip, and
    render integer-valued floats without a trailing '.0'."""
    if pd.isna(v):
        return ""
    if isinstance(v, float) and float(v).is_integer():
        return str(int(v))
    return str(v).strip()


def _read_raw(content: bytes, filename: str) -> pd.DataFrame:
    """Read the raw bytes of one upload into a DataFrame, no normalization.
    The new export is a single-header sheet (named 'Sheet'); fall back to the
    first sheet for workbooks and a plain read for .csv.

    ExtPONumber is forced to text so long numeric PO numbers aren't coerced to
    floats (which would lose precision / show scientific notation). Unknown
    dtype keys are ignored by pandas when the column isn't present."""
    suffix = Path(filename).suffix.lower()
    buf = io.BytesIO(content)
    str_cols = {"ExtPONumber": str, "Ext PO #": str}
    if suffix in (".xlsx", ".xlsm", ".xls"):
        try:
            return pd.read_excel(buf, sheet_name="Sheet", dtype=str_cols)
        except ValueError:
            buf.seek(0)
            return pd.read_excel(buf, dtype=str_cols)
    if suffix == ".csv":
        return pd.read_csv(buf, dtype=str_cols)
    raise ValueError(f"Unsupported file type for {filename!r}: {suffix}")


def _to_cancelled(v) -> str:
    """Map the new boolean IsCancelled to the pipeline's 'Yes'/'' convention.
    Handles real booleans (xlsx) and 'True'/'False' strings (csv)."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    if isinstance(v, str):
        return "Yes" if v.strip().lower() in ("true", "yes", "1") else ""
    return "Yes" if bool(v) else ""


def _normalize_new_export(df: pd.DataFrame, filename: str) -> pd.DataFrame:
    """Translate a new-export DataFrame into the internal pipeline schema.

    Produces exactly the columns the proven pipeline expects:
        Company, PO #, Adjustment Date, Vendor, Team/Performer,
        Ticket Cost Total Start, Ticket Cost Total End, Total Adjustment,
        Cancelled, User  (+ Remove if the manual flag column is present)

    Key derivations:
      * Total Adjustment = End - Start  (TicketCostTotal - InitialTicketCostTotal)
      * Adjustment Date  = AdjustedDateTimeUTC, UTC → US Central, date only
      * Cancelled        = "Yes" when IsCancelled is truthy, else ""
    """
    missing = NEW_REQUIRED - set(df.columns)
    if missing:
        raise ValueError(
            f"{filename!r} doesn't look like the new PO Cost Changes export — "
            f"missing columns: {sorted(missing)}. Found: {list(df.columns)}"
        )

    out = pd.DataFrame()
    for src, dst in NEW_COLUMN_MAP.items():
        out[dst] = df[src]

    # Money: coerce to numeric, then derive the adjustment (End - Start).
    start = pd.to_numeric(out["Ticket Cost Total Start"], errors="coerce")
    end = pd.to_numeric(out["Ticket Cost Total End"], errors="coerce")
    out["Ticket Cost Total Start"] = start
    out["Ticket Cost Total End"] = end
    out["Total Adjustment"] = end.fillna(0) - start.fillna(0)

    # Cancelled flag → pipeline convention.
    out["Cancelled"] = df["IsCancelled"].map(_to_cancelled).astype("string")

    # Adjustment Date — override to the filename's date for every row in the
    # file. Falls back to the UTC → US Central conversion only when the filename
    # has no parseable YYYY-MM-DD date.
    override = _filename_adjustment_date(filename)
    if override is not None:
        out["Adjustment Date"] = override
    else:
        # Fallback: UTC datetime → US Central → date only. tz_localize(None)
        # drops the tz so downstream comparisons / Excel writing get naive
        # datetimes.
        ts = pd.to_datetime(df["AdjustedDateTimeUTC"], errors="coerce", utc=True)
        out["Adjustment Date"] = (
            ts.dt.tz_convert(LOCAL_TZ).dt.tz_localize(None).dt.normalize()
        )

    # AccountEmail + ExtPONumber are aggregation keys (and feed the memo).
    # Clean to plain strings with "" for blanks so grouping is consistent.
    out["AccountEmail"] = (
        df["AccountEmail"].map(_clean_key_str) if "AccountEmail" in df.columns else ""
    )
    out["ExtPONumber"] = (
        df["ExtPONumber"].map(_clean_key_str) if "ExtPONumber" in df.columns else ""
    )

    # Same-day exclusion flag: exclude rows where the cost change happened on
    # the same calendar day the PO was created. Both AdjustedDateTimeUTC and
    # CreatedDate are UTC; we compare their dates in US Central. A blank
    # CreatedDate is never a match (those rows are kept). Computed here from the
    # raw timestamps; carried as a hidden helper (dropped before any sheet is
    # written) and folded into the exclusion mask in process_files.
    adj_ct = pd.to_datetime(df["AdjustedDateTimeUTC"], errors="coerce", utc=True).dt.tz_convert(LOCAL_TZ)
    if "CreatedDate" in df.columns:
        cre_ct = pd.to_datetime(df["CreatedDate"], errors="coerce", utc=True).dt.tz_convert(LOCAL_TZ)
        same_day = (
            adj_ct.dt.normalize().eq(cre_ct.dt.normalize())
            & adj_ct.notna()
            & cre_ct.notna()
        )
        # CreatedDate is displayed (Source Data / Excluded) as a US-Central
        # date with no timestamp. tz_localize(None) + normalize() yields a
        # midnight-naive datetime that renders mm/dd/yyyy via _write_sheet.
        out["CreatedDate"] = cre_ct.dt.tz_localize(None).dt.normalize()
    else:
        same_day = pd.Series(False, index=out.index)
        out["CreatedDate"] = pd.NaT
    out[EXCLUDE_SAME_DATE_COL] = same_day.fillna(False).to_numpy()

    # Preserve the manual "Remove" flag column if the user added one.
    remove_col = next(
        (c for c in df.columns if str(c).strip().lower() == "remove"), None
    )
    if remove_col is not None:
        out["Remove"] = df[remove_col]

    # Carry through extra source columns unchanged (kept for Source Data and
    # any downstream rules). Silently skip any that aren't present.
    for col in PASSTHROUGH_COLUMNS:
        if col in df.columns:
            out[col] = df[col]

    return out


SOURCE_VIEW_ORDER = [
    "Company", "PO #", "Vendor", "Team/Performer",
    "Ticket Cost Total Start", "Ticket Cost Total End", "User",
    "Total Adjustment", "Cancelled", "Adjustment Date",
    "AccountEmail", "ExtPONumber", "CreatedDate",
]


def _order_source_view(df: pd.DataFrame) -> pd.DataFrame:
    """Put the internal-schema columns in the familiar Source Data order, with
    any extras (e.g. VenueName, Remove) kept at the end and helper columns
    dropped."""
    front = [c for c in SOURCE_VIEW_ORDER if c in df.columns]
    rest = [c for c in df.columns if c not in front and not str(c).startswith("_")]
    return df[front + rest]


def convert_new_format(file_bytes: bytes, filename: str = "") -> bytes:
    """Zone 1 converter: read a raw PO Cost Changes export and return a single
    'Source Data' sheet of the normalized, cleaner data — the same internal
    schema shown on the Source Data tab of the full outputs.

    The user can review/edit this file (fix a vendor or cost, add a 'Remove'
    column with an X, delete rows) and then upload it into Zone 2, which reads
    the 'Source Data' tab via the re-upload path. A bad file raises ValueError,
    surfaced to the caller.
    """
    df = _read_one(file_bytes, filename)

    # Drop the same-day helper if present, apply the Cancelled override so the
    # Total Adjustment matches the Source Data tab, then order the columns.
    df = df.drop(columns=[EXCLUDE_SAME_DATE_COL], errors="ignore")
    view = _order_source_view(_apply_cancelled_override_raw(df)).reset_index(drop=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, "Source Data", view)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _read_one(content: bytes, filename: str) -> pd.DataFrame:
    """Read one uploaded file and normalize it to the internal schema.

    Two accepted shapes:
      * a fresh TicketVault PO Cost Changes export (single 'Sheet'), or
      * one of THIS app's own output workbooks (or a Zone 1 converted file) —
        detected by a 'Source Data' tab — re-uploaded after the user hand-edited
        that tab. In that case we re-ingest the edited Source Data rows.
    """
    suffix = Path(filename).suffix.lower()
    if suffix in (".xlsx", ".xlsm", ".xls"):
        try:
            xls = pd.ExcelFile(io.BytesIO(content))
            sheets = set(xls.sheet_names)
        except Exception:
            sheets = set()
        if "Source Data" in sheets and "Sheet" not in sheets:
            sd = pd.read_excel(xls, sheet_name="Source Data", dtype={"ExtPONumber": str})
            return _normalize_reupload(sd, filename)

    raw = _read_raw(content, filename)
    # A Zone 1 "Modified" file: rename its columns back to the raw export names
    # and process exactly like the raw export (it's the same data, reformatted).
    if _is_modified_format(raw):
        raw = raw.rename(columns=MODIFIED_TO_RAW)
    return _normalize_new_export(raw, filename)


def _file_has_remove_column(content: bytes, filename: str) -> bool:
    """True if one uploaded file contains a 'Remove' column (case-insensitive,
    any position). Reads only the header of the same sheet `_read_one` would
    process, so detection matches how the file is actually ingested. On any read
    error, returns True (don't block on Remove grounds — let normal processing
    surface the real error)."""
    suffix = Path(filename).suffix.lower()
    try:
        if suffix in (".xlsx", ".xlsm", ".xls"):
            xls = pd.ExcelFile(io.BytesIO(content))
            sheets = set(xls.sheet_names)
            if "Source Data" in sheets and "Sheet" not in sheets:
                target = "Source Data"
            elif "Sheet" in sheets:
                target = "Sheet"
            else:
                target = xls.sheet_names[0]
            cols = pd.read_excel(xls, sheet_name=target, nrows=0).columns
        else:
            cols = pd.read_csv(io.BytesIO(content), nrows=0).columns
    except Exception:
        return True
    return any(str(c).strip().lower() == "remove" for c in cols)


def files_missing_remove_column(file_list) -> list:
    """Given a list of (content, filename), return the filenames that do NOT
    contain a 'Remove' column. Used to hard-block processing until every
    uploaded file has the column."""
    return [fn for (content, fn) in file_list if not _file_has_remove_column(content, fn)]


# Internal-schema columns that must be present to re-ingest an output file's
# "Source Data" tab. Everything else is derived or optional.
REUPLOAD_REQUIRED = {
    "Company", "PO #", "Vendor", "Team/Performer",
    "Ticket Cost Total Start", "Ticket Cost Total End",
}


def _normalize_reupload(df: pd.DataFrame, filename: str) -> pd.DataFrame:
    """Re-ingest a 'Source Data' tab (from a full output workbook or a Zone 1
    converted file). The tab is already in the internal schema, so columns pass
    straight through and only the derived bits are recomputed:
      * Total Adjustment = End − Start  (so manual cost edits take effect)
      * the same-day exclusion flag, from Adjustment Date vs CreatedDate

    Adjustment Date is taken as-is from the sheet. A 'Remove' column is honored
    if the user added one.
    """
    df = df.rename(columns={c: str(c).strip() for c in df.columns})

    missing = REUPLOAD_REQUIRED - set(df.columns)
    if missing:
        raise ValueError(
            f"{filename!r} looks like a PO Cost Changes output/converted file, but "
            f"its 'Source Data' tab is missing columns: {sorted(missing)}."
        )

    out = pd.DataFrame()
    for col in ["Company", "PO #", "Vendor", "Team/Performer", "User"]:
        if col in df.columns:
            out[col] = df[col]

    start = pd.to_numeric(df["Ticket Cost Total Start"], errors="coerce")
    end = pd.to_numeric(df["Ticket Cost Total End"], errors="coerce")
    out["Ticket Cost Total Start"] = start
    out["Ticket Cost Total End"] = end
    out["Total Adjustment"] = end.fillna(0) - start.fillna(0)

    if "Cancelled" in df.columns:
        out["Cancelled"] = df["Cancelled"].map(_to_cancelled).astype("string")
    else:
        out["Cancelled"] = pd.Series([""] * len(df), dtype="string")

    out["Adjustment Date"] = pd.to_datetime(df.get("Adjustment Date"), errors="coerce")

    out["AccountEmail"] = df["AccountEmail"].map(_clean_key_str) if "AccountEmail" in df.columns else ""
    out["ExtPONumber"] = df["ExtPONumber"].map(_clean_key_str) if "ExtPONumber" in df.columns else ""

    if "CreatedDate" in df.columns:
        out["CreatedDate"] = pd.to_datetime(df["CreatedDate"], errors="coerce").dt.normalize()
    else:
        out["CreatedDate"] = pd.NaT

    # Same-day exclusion — recompute from Adjustment Date vs CreatedDate at the
    # date level (both already US Central; blank CreatedDate never matches).
    adj = pd.to_datetime(out["Adjustment Date"], errors="coerce").dt.normalize()
    cre = out["CreatedDate"]
    same_day = adj.eq(cre) & adj.notna() & cre.notna()
    out[EXCLUDE_SAME_DATE_COL] = same_day.fillna(False).to_numpy()

    remove_col = next((c for c in df.columns if str(c).strip().lower() == "remove"), None)
    if remove_col is not None:
        out["Remove"] = df[remove_col]

    for col in PASSTHROUGH_COLUMNS:
        if col in df.columns:
            out[col] = df[col]

    return out


def _write_sheet(wb, sheet_name: str, df: pd.DataFrame) -> None:
    """Write df to a new sheet with the same styling as the reference project:
    blue header, alternating row fills, borders, frozen header, auto-filter,
    currency formatting on money columns, mm/dd/yyyy on Adjustment Date, and a
    yellow highlight on any "User" cell that is a known accounting-team user.

    Sheet names are sanitized to Excel's rules (≤31 chars, no `:\\/?*[]`).
    """
    safe = sheet_name[:31]
    for ch in r":\/?*[]":
        safe = safe.replace(ch, "_")

    ws = wb.create_sheet(safe)
    cols = list(df.columns)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    fill_odd  = PatternFill("solid", start_color="FFFFFF")
    fill_even = PatternFill("solid", start_color="EEF2FF")
    money_cols = {"Total Start", "Total End", "Total Adjustment", "Total", "Total Cost"}

    for ri, row in enumerate(df.itertuples(index=False), 2):
        row_fill = fill_even if ri % 2 == 0 else fill_odd
        for ci, val in enumerate(row, 1):
            col_name = cols[ci - 1]
            # Cast pandas NA / NaT to None so openpyxl writes empty cells.
            if pd.isna(val):
                val = None
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
            cell.fill = row_fill
            # Flag accounting-team users (overrides the alternating row fill).
            # Only sheets carrying a "User" column (Source Data / Excluded) are
            # affected; everything else is untouched.
            if col_name == "User" and _is_accounting_user(val):
                cell.fill = ACCOUNTING_FILL
            if col_name in ("Adjustment Date", "Date", "CreatedDate", "PO Created") and val is not None:
                cell.number_format = "mm/dd/yyyy"
            elif col_name in money_cols and val is not None:
                cell.number_format = '"$"#,##0.00;[Red]"-$"#,##0.00'

    # Column widths sized to content (capped at 55).
    for ci, col in enumerate(cols, 1):
        max_len = len(str(col))
        for row in df.itertuples(index=False):
            v = row[ci - 1]
            max_len = max(max_len, 0 if v is None or (isinstance(v, float) and pd.isna(v)) else len(str(v)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 55)

    ws.freeze_panes = "A2"
    # Filter dropdowns on the header row of every sheet (whole used range, or
    # just the header when there are no data rows).
    ws.auto_filter.ref = ws.dimensions


def _created_paren(row) -> str:
    """The '(PO created date …)' suffix listing every distinct created date in
    the aggregated group (comma-separated, mm/dd/yyyy). Empty string when there
    is no created date."""
    cds = row.get("CreatedDate") if hasattr(row, "get") else None
    if isinstance(cds, tuple):
        dates = [d for d in cds if not pd.isna(d)]
    elif cds is None or (not isinstance(cds, tuple) and pd.isna(cds)):
        dates = []
    else:
        dates = [cds]
    if not dates:
        return ""
    joined = ", ".join(pd.Timestamp(d).strftime("%m/%d/%Y") for d in dates)
    return f"(PO created date {joined})"


def _memo_parts(row) -> list[str]:
    """The non-empty leading memo fields, in order: performer, account email,
    ext PO #. Blanks are dropped so joins never produce empty segments."""
    team = _clean_key_str(row["Team/Performer"])
    email = _clean_key_str(row["AccountEmail"])
    extpo = _clean_key_str(row["ExtPONumber"])
    return [p for p in (team, email, extpo) if p]


def _full_memo(row) -> str:
    """Full memo: 'Performer / email / ext PO # / Cost Changes (Company) (PO
    created date …)'. Empty fields are skipped; the created-date suffix is
    omitted when there's no created date."""
    orig = _clean_key_str(row["Original Company"])
    parts = _memo_parts(row)
    parts.append(f"Cost Changes ({orig})")
    memo = " / ".join(parts)
    cp = _created_paren(row)
    return f"{memo} {cp}" if cp else memo


def _memo2(row) -> str:
    """Memo2 (PD-format bills): 'Performer / email / ext PO # (PO created date
    …)' — no company, no 'Cost Changes' text."""
    memo = " / ".join(_memo_parts(row))
    cp = _created_paren(row)
    if not memo:
        return cp
    return f"{memo} {cp}" if cp else memo


def _build_bills_and_expenses(
    cleaned: pd.DataFrame,
    tc_vendors: set[str] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Build the Bills and Expenses ledger views.

    Bills: one row per PO whose Total Adjustment > 0. Total stays positive.

    Expenses: two rows per PO whose Total Adjustment < 0.
      - Line A: Category = 'Inventory Asset', Total = Total Adjustment (negative)
      - Line B: Category = '<Vendor> (TC)' if the final vendor is on the TC list
                (third tab of Vendors_Open.xlsx), otherwise
                'Due from Vendors - Open'. Total = -Total Adjustment (positive),
                so the pair sums to zero.

    Both share the same column shape and a single global 'Expense #' counter
    incrementing across the whole result (Bills + Expenses interleaved by
    display order, then by PO #).
    """
    if tc_vendors is None:
        tc_vendors = get_tc_vendors()

    ledger_cols = [
        "Company", "Account", "Date", "Category", "Expense #",
        "Vendor", "Memo", "Description", "Total",
    ]

    # Empty input → empty frames with the right shape, so downstream writers
    # can still render headers.
    if cleaned.empty:
        empty = pd.DataFrame(columns=ledger_cols)
        return empty.copy(), empty.copy()

    # Order events by display order, then date and vendor, so the global
    # Expense # is deterministic and groups each company's rows together.
    df = cleaned.copy()
    df["_display_order"] = df["Company"].map(_sort_key)
    df = df.sort_values(
        ["_display_order", "Adjustment Date", "Vendor", "Team/Performer"],
        kind="mergesort",
    ).reset_index(drop=True)
    df["Expense #"] = range(1, len(df) + 1)

    df["_memo"] = df.apply(_full_memo, axis=1)

    # Bills: positive adjustments → single row each.
    bills_src = df[df["Total Adjustment"] > 0].copy()
    bills = pd.DataFrame({
        "Company":     bills_src["Original Company"].map(company_value),
        "Account":     "Clearing Account",
        "Date":        bills_src["Adjustment Date"],
        "Category":    "Inventory Asset",
        "Expense #":   bills_src["Expense #"],
        "Vendor":      bills_src["Vendor"],
        "Memo":        bills_src["_memo"],
        "Description": bills_src["_memo"],
        "Total":       bills_src["Total Adjustment"],
    })

    # Expenses: negative adjustments → two rows each, summing to zero.
    exp_src = df[df["Total Adjustment"] < 0].copy()
    if not exp_src.empty:
        # Line A — Inventory Asset, negative
        line_a = pd.DataFrame({
            "Company":     exp_src["Original Company"].map(company_value),
            "Account":     "Clearing Account",
            "Date":        exp_src["Adjustment Date"],
            "Category":    "Inventory Asset",
            "Expense #":   exp_src["Expense #"],
            "Vendor":      exp_src["Vendor"],
            "Memo":        exp_src["_memo"],
            "Description": exp_src["_memo"],
            "Total":       exp_src["Total Adjustment"],   # already negative
        })
        # Line B — Vendor (TC) or Due from Vendors - Open, positive offset
        line_b = pd.DataFrame({
            "Company":     exp_src["Original Company"].map(company_value),
            "Account":     "Clearing Account",
            "Date":        exp_src["Adjustment Date"],
            "Category":    exp_src["Vendor"].map(lambda v: offset_category(v, tc_vendors)),
            "Expense #":   exp_src["Expense #"],
            "Vendor":      exp_src["Vendor"],
            "Memo":        exp_src["_memo"],
            "Description": exp_src["_memo"],
            "Total":       -exp_src["Total Adjustment"],  # positive (flips sign)
        })
        # Interleave A,B,A,B,... by sorting on (Expense #, line_order)
        line_a["_line"] = 0
        line_b["_line"] = 1
        expenses = (
            pd.concat([line_a, line_b], ignore_index=True)
            .sort_values(["Expense #", "_line"], kind="mergesort")
            .drop(columns=["_line"])
            .reset_index(drop=True)
        )
    else:
        expenses = pd.DataFrame(columns=ledger_cols)

    # Attach the short sheet/file label as a hidden helper column so the
    # per-company tab/file filtering can use it. Map via Expense #.
    display_by_expense = dict(zip(df["Expense #"], df["Company"].map(file_label)))
    if not expenses.empty:
        expenses["_display_label"] = expenses["Expense #"].map(display_by_expense)
    if not bills.empty:
        bills["_display_label"] = bills["Expense #"].map(display_by_expense)

    return bills[ledger_cols + (["_display_label"] if "_display_label" in bills else [])], \
           expenses[ledger_cols + (["_display_label"] if "_display_label" in expenses else [])]


def _apply_cancelled_override_raw(df: pd.DataFrame) -> pd.DataFrame:
    """Apply the Cancelled override to a RAW DataFrame (pre-pipeline shape).

    When Cancelled == "Yes", set Total Adjustment = -(Ticket Cost Total End).
    This mirrors step 8 of transform(), but for raw rows so the Source Data
    and Excluded tabs show numbers that reconcile cleanly to the Combined
    ledger (which uses the same override internally).

    Returns a new DataFrame with the column updated; the original is untouched.
    """
    if "Cancelled" not in df.columns or "Total Adjustment" not in df.columns:
        return df
    out = df.copy()
    is_cancelled = out["Cancelled"].astype("string").str.strip().str.lower().eq("yes")
    if "Ticket Cost Total End" in out.columns:
        end = pd.to_numeric(out["Ticket Cost Total End"], errors="coerce")
        out.loc[is_cancelled, "Total Adjustment"] = -end[is_cancelled]
    return out


def _write_summary_sheet(wb, combined_ledger: pd.DataFrame) -> None:
    """Write the 'Summary' sheet: a pivot-style outline view of the Combined
    ledger, grouped Company > Vendor > Description, with one row per (Company,
    Vendor, Description, Date). Subtotal rows after each Company group; grand
    total at the bottom. Outline groups are expanded by default; the user can
    collapse with Excel's outline +/− buttons in the left margin.

    Matches the layout in the user's screenshot: blue header band, currency
    formatting on Total (red for negatives), date in mm/dd/yyyy.
    """
    ws = wb.create_sheet("Summary")

    # ── Shared styling ──────────────────────────────────────────────────────
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    company_font = Font(name="Arial", bold=True, size=10)
    company_fill = PatternFill("solid", start_color="D9E2F3")
    # Subtotal rows: black bold font. Negative subtotals still render red via the
    # number format's [Red] section; positives stay black (not red).
    subtotal_font = Font(name="Arial", bold=True, color="000000", size=10)
    subtotal_fill = PatternFill("solid", start_color="EEF2FF")
    grand_font = Font(name="Arial", bold=True, size=11)
    body_font = Font(name="Arial", size=10)
    money_fmt = '"$"#,##0.00;[Red]"-$"#,##0.00'

    # ── Empty data → just write headers and bail ────────────────────────────
    if combined_ledger.empty:
        for ci, h in enumerate(["Company", "Vendor", "Description", "Date", "Total"], 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = header_align; cell.border = border
        ws.freeze_panes = "A2"
        return

    # ── Aggregate to (Company, Vendor, Description, Date) ──────────────────
    df = combined_ledger.copy()
    df["Total"] = pd.to_numeric(df["Total"], errors="coerce").fillna(0.0)
    grouped = (
        df.groupby(["Company", "Vendor", "Description", "Date"], dropna=False, as_index=False)
        ["Total"].sum()
    )
    # Sort: company by the same DISPLAY_ORDER used everywhere else; everything
    # else alphabetically/chronologically.
    grouped["_co_order"] = grouped["Company"].map(_sort_key)
    grouped = grouped.sort_values(
        ["_co_order", "Vendor", "Description", "Date"], kind="mergesort"
    ).drop(columns=["_co_order"]).reset_index(drop=True)

    # ── Title + header row ──────────────────────────────────────────────────
    # Row 1 — "Sum of Total" label (matches the screenshot's banner)
    ws.cell(row=1, column=1, value="Sum of Total").font = Font(name="Arial", bold=True, size=11)
    # Row 2 — column headers
    headers = ["Company", "Vendor", "Description", "Date", "Total"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = header_align; cell.border = border

    # ── Body: one row per (Company, Vendor, Description, Date) with
    #    Company shown only on the first row of its group + a subtotal row. ──
    row = 3
    company_row_ranges: list[tuple[str, int, int]] = []  # (company, start, end_data)
    for company, co_grp in grouped.groupby("Company", sort=False, dropna=False):
        co_start = row
        first_row_of_company = True
        for _, r in co_grp.iterrows():
            # Company label — only on the first row of the group
            ws.cell(row=row, column=1, value=str(company) if first_row_of_company else None)
            ws.cell(row=row, column=2, value=str(r["Vendor"]) if pd.notna(r["Vendor"]) else None)
            ws.cell(row=row, column=3, value=str(r["Description"]) if pd.notna(r["Description"]) else None)
            date_val = r["Date"] if pd.notna(r["Date"]) else None
            ws.cell(row=row, column=4, value=date_val)
            ws.cell(row=row, column=5, value=float(r["Total"]))

            # Styling
            ws.cell(row=row, column=4).number_format = "mm/dd/yyyy"
            ws.cell(row=row, column=5).number_format = money_fmt
            if first_row_of_company:
                ws.cell(row=row, column=1).font = company_font
                ws.cell(row=row, column=1).fill = company_fill
            for ci in range(1, 6):
                cell = ws.cell(row=row, column=ci)
                if cell.font.name is None:
                    cell.font = body_font
                cell.border = border
                cell.alignment = Alignment(vertical="center")

            first_row_of_company = False
            row += 1

        # Subtotal row for this Company
        ws.cell(row=row, column=1, value=f"{company} Total")
        ws.cell(row=row, column=5, value=f"=SUBTOTAL(9,E{co_start}:E{row - 1})")
        for ci in range(1, 6):
            cell = ws.cell(row=row, column=ci)
            cell.font = subtotal_font; cell.fill = subtotal_fill
            cell.border = border; cell.alignment = Alignment(vertical="center")
        ws.cell(row=row, column=5).number_format = money_fmt

        company_row_ranges.append((str(company), co_start, row - 1))
        row += 1

    # ── Grand total at the bottom ──────────────────────────────────────────
    grand_row = row
    ws.cell(row=grand_row, column=1, value="Grand Total")
    if company_row_ranges:
        first_start = company_row_ranges[0][1]
        last_end = company_row_ranges[-1][2]
        # Use SUM (not SUBTOTAL) since we want it to ignore the subtotal rows.
        # Easiest way: sum the per-company subtotal cells.
        subtotal_cells = [f"E{end + 1}" for _, _, end in company_row_ranges]
        ws.cell(row=grand_row, column=5, value=f"={'+'.join(subtotal_cells)}")
    else:
        ws.cell(row=grand_row, column=5, value=0)
    for ci in range(1, 6):
        cell = ws.cell(row=grand_row, column=ci)
        cell.font = grand_font; cell.fill = company_fill
        cell.border = Border(left=thin, right=thin,
                             top=Side(style="medium", color="000000"), bottom=thin)
        cell.alignment = Alignment(vertical="center")
    ws.cell(row=grand_row, column=5).number_format = money_fmt

    # ── Outline groups for collapse/expand on Company column ───────────────
    # Excel outline: data rows of each company are grouped one level deep;
    # collapsing hides everything except the subtotal row. Default expanded.
    for _, start, end in company_row_ranges:
        if end >= start:
            for r in range(start, end + 1):
                ws.row_dimensions[r].outline_level = 1
                ws.row_dimensions[r].hidden = False
    ws.sheet_properties.outlinePr.summaryBelow = True

    # ── Column widths and freeze pane ──────────────────────────────────────
    widths = {1: 24, 2: 24, 3: 60, 4: 12, 5: 14}
    for ci, w in widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A3"


def _combined_ledger(bills_df: pd.DataFrame, expenses_df: pd.DataFrame) -> pd.DataFrame:
    """The 'Combined' ledger: bills as-is plus the negative (Inventory Asset)
    leg of each expense pair, ordered by Expense #. Used for the Combined tab
    and the Summary pivot."""
    drop_helper = lambda d: d.drop(columns=[c for c in ["_display_label"] if c in d.columns])
    bills_visible = drop_helper(bills_df)
    expenses_visible = drop_helper(expenses_df)
    if not expenses_visible.empty:
        expense_singles = expenses_visible[expenses_visible["Category"] == "Inventory Asset"]
    else:
        expense_singles = expenses_visible
    return (
        pd.concat([bills_visible, expense_singles], ignore_index=True)
        .sort_values("Expense #", kind="mergesort")
        .reset_index(drop=True)
    )


def _build_combined_workbook(
    source_df: pd.DataFrame,
    bills_df: pd.DataFrame,
    expenses_df: pd.DataFrame,
    all_company_labels: list[str],
    excluded_df: pd.DataFrame | None = None,
    pd_bills_df: pd.DataFrame | None = None,
) -> bytes:
    """Build the multi-sheet combined workbook:
       - 'Source Data' — the raw merged upload, untouched
       - 'Combined'    — one row per aggregated event (ledger format)
       - 'Bills'       — positive-adjustment events (one row each)
       - One tab per company — the company's Expense pairs (debit/credit)
       - 'Excluded'    — raw rows filtered out via the Remove=X flag
                         (last tab; only shown if any rows were excluded)
       - Empty per-company tabs get a red tab color
    """
    drop_helper = lambda d: d.drop(columns=[c for c in ["_display_label"] if c in d.columns])
    bills_visible = drop_helper(bills_df)
    expenses_visible = drop_helper(expenses_df)

    # 'Combined' = every aggregated event as a single ledger row.
    combined_ledger = _combined_ledger(bills_df, expenses_df)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Summary first — pivot-style outline of the Combined ledger.
    _write_summary_sheet(wb, combined_ledger)

    # Source data (full; drop the per-company label helper column)
    _write_sheet(wb, "Source Data",
                 source_df.drop(columns=[c for c in ["_label"] if c in source_df.columns]))

    # Excluded tab — right after Source Data so it's adjacent to the raw view.
    # Only added when at least one row got flagged Remove=X. Shape matches
    # Source Data so the user sees exactly which raw rows were removed.
    if excluded_df is not None and len(excluded_df) > 0:
        _write_sheet(wb, "Excluded",
                     excluded_df.drop(columns=[c for c in ["_label"] if c in excluded_df.columns])
                     .reset_index(drop=True))

    # Combined ledger
    _write_sheet(wb, "Combined", combined_ledger)

    # Bills tab — same PD/QBO layout as the individual per-company bills files
    # (Company · Bill No. · PO Created · Account · Vendor · Memo2 ·
    # Team/Performer · Memo · Total Cost · Seasons), all companies together.
    if pd_bills_df is not None:
        bills_tab = pd_bills_df.drop(
            columns=[c for c in ["_display_label"] if c in pd_bills_df.columns]
        )[PD_BILLS_COLUMNS].reset_index(drop=True)
    else:
        bills_tab = bills_visible
    _write_sheet(wb, "Bills", bills_tab)

    # One tab per company with Expense pairs
    for label in all_company_labels:
        if "_display_label" in expenses_df.columns:
            tab_df = (
                expenses_df[expenses_df["_display_label"] == label]
                .drop(columns=["_display_label"])
                .reset_index(drop=True)
            )
        else:
            tab_df = expenses_df.iloc[0:0]
        _write_sheet(wb, label, tab_df)
        if len(tab_df) == 0:
            wb[label[:31]].sheet_properties.tabColor = "FF0000"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _company_views(
    source_view: pd.DataFrame | None,
    excluded_view: pd.DataFrame | None,
    label: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Return (source_rows, excluded_rows) for one company, with the hidden
    '_label' helper column removed. Empty (header-only) frames are returned
    when there's nothing for that company."""
    def _filt(df: pd.DataFrame | None) -> pd.DataFrame:
        if df is None:
            return pd.DataFrame()
        keep = [c for c in df.columns if c != "_label"]
        if "_label" not in df.columns:
            return df[keep].reset_index(drop=True)
        return df[df["_label"] == label][keep].reset_index(drop=True)
    return _filt(source_view), _filt(excluded_view)


def _build_company_file(
    first_sheet_name: str,
    first_df: pd.DataFrame,
    label: str,
    source_view: pd.DataFrame | None = None,
    excluded_view: pd.DataFrame | None = None,
    summary_ledger: pd.DataFrame | None = None,
) -> bytes:
    """Per-company download file. Tab 1 = data sheet (Expenses or Bills); tab 2
    = that company's Summary pivot; tabs 3/4 = its own Source Data and
    Excluded."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_sheet(wb, first_sheet_name, first_df.reset_index(drop=True))
    _write_summary_sheet(wb, summary_ledger if summary_ledger is not None else pd.DataFrame())

    src_rows, exc_rows = _company_views(source_view, excluded_view, label)
    _write_sheet(wb, "Source Data", src_rows)
    _write_sheet(wb, "Excluded", exc_rows)
    if len(exc_rows) == 0:
        wb["Excluded"].sheet_properties.tabColor = "FF0000"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_company_workbook(
    label: str,
    expenses_df: pd.DataFrame,
    source_view: pd.DataFrame | None = None,
    excluded_view: pd.DataFrame | None = None,
    summary_ledger: pd.DataFrame | None = None,
) -> bytes:
    """Per-company Expenses file: Expenses + Summary + Source Data + Excluded."""
    if "_display_label" in expenses_df.columns:
        e = expenses_df[expenses_df["_display_label"] == label].drop(columns=["_display_label"])
    else:
        e = expenses_df.iloc[0:0]
    return _build_company_file("Expenses", e, label, source_view, excluded_view, summary_ledger)


def _write_single_sheet_xlsx(df: pd.DataFrame, sheet_name: str = "PO Cost Changes") -> bytes:
    """Single-sheet styled workbook — kept for compatibility / future use."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, sheet_name, df)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Column order for the per-company PD-format bills files (matches the Purchase
# Details output layout exactly).
PD_BILLS_COLUMNS = [
    "Company", "Bill No.", "PO Created", "Account", "Vendor",
    "Memo2", "Team/Performer", "Memo", "Total Cost", "Seasons",
]


def build_pd_bills(cleaned: pd.DataFrame) -> pd.DataFrame:
    """Build the per-company "bills" rows in the Purchase Details output layout.

    Source rows are the positive cost adjustments (the items that land on the
    Bills tab). Columns:
      Company        — raw/original company (matches the memo's company suffix)
      Bill No.       — random 8-digit integer, one per row
      PO Created     — the Adjustment Date (the filename's date)
      Account        — "Inventory Asset"
      Vendor         — renamed vendor
      Memo2          — "Performer / email / ext PO #" (no company)
      Team/Performer — full memo "… / Cost Changes (Company)"
      Memo           — same as Team/Performer
      Total Cost     — the positive adjustment amount
      Seasons        — blank (manual entry by the user)

    Includes a hidden "_display_label" column so process_files can split the
    rows into one file per QBO company; it's dropped before writing.
    """
    if cleaned.empty:
        out = pd.DataFrame(columns=PD_BILLS_COLUMNS + ["_display_label"])
        return out

    src = cleaned[cleaned["Total Adjustment"] > 0].copy()
    if src.empty:
        return pd.DataFrame(columns=PD_BILLS_COLUMNS + ["_display_label"])

    full = src.apply(_full_memo, axis=1)
    out = pd.DataFrame({
        "Company":        src["Original Company"].map(company_value),
        "Bill No.":       [random.randint(10000000, 99999999) for _ in range(len(src))],
        "PO Created":     src["Adjustment Date"],
        "Account":        "Inventory Asset",
        "Vendor":         src["Vendor"],
        "Memo2":          src.apply(_memo2, axis=1),
        "Team/Performer": full,
        "Memo":           full,
        "Total Cost":     src["Total Adjustment"],
        "Seasons":        "",
        "_display_label": src["Company"].map(file_label),   # short label for per-company split
    })
    # Deterministic order: by company, then date, vendor, memo.
    out = out.sort_values(
        ["_display_label", "PO Created", "Vendor", "Memo2"], kind="mergesort"
    ).reset_index(drop=True)
    return out


def process_files(
    file_list: list[tuple[bytes, str]],
) -> dict:
    """Process N uploaded files into the bundle the Flask app needs.

    Args:
        file_list: PO Cost Changes uploads — list of (bytes, filename) tuples.

    Row exclusion is driven solely by the manual "Remove" = X flag; the new
    export's PurchaseDetailMatchFound column is ignored per spec.

    Returns:
        {
          "date_range": "May 1st thru May 3rd 2026",
          "combined": <xlsx bytes>,                        # multi-tab workbook
          "companies": {label: <xlsx bytes>, ...},          # only labels with data
          "all_companies": [...],                           # all QBO labels, sorted
          "stats": { "Combined": {...}, "<label>": {...}, ... },
          "dropped": {"unmapped_companies": {...}, "total_dropped_rows": N},
          "excluded": {"po_count": N, "row_count": N},      # Remove=X rows
        }
    """
    if not file_list:
        raise ValueError("No files provided")

    # 1) Read + merge all PO Cost Changes uploads.
    frames = []
    for content, filename in file_list:
        df = _read_one(content, filename)
        frames.append(df)
        log.info("Read %s: %d rows", filename, len(df))
    merged = pd.concat(frames, ignore_index=True, sort=False)

    # 1a) Filter out any company that isn't in the QBO master mapping.
    #     Doing this on the raw merge means Source Data and Excluded tabs
    #     only show in-scope rows, and the math reconciles cleanly:
    #     Source Data total − Excluded total = Combined total.
    mapping_now = get_mapping()
    valid_lower = set(mapping_now.keys()) | {v.lower() for v in mapping_now.values()}
    raw_company = merged["Company"].astype("string").str.strip().str.lower()
    in_scope_mask = raw_company.isin(valid_lower)
    out_of_scope_counts: dict[str, int] = {}
    if (~in_scope_mask).any():
        out_of_scope = merged.loc[~in_scope_mask, "Company"].astype("string").fillna("(blank)")
        out_of_scope_counts = out_of_scope.value_counts().to_dict()
        log.info(
            "Ignored %d rows from %d out-of-scope compan%s: %s",
            int((~in_scope_mask).sum()),
            len(out_of_scope_counts),
            "y" if len(out_of_scope_counts) == 1 else "ies",
            ", ".join(f"{k} ({v})" for k, v in out_of_scope_counts.items()),
        )
    merged = merged[in_scope_mask].reset_index(drop=True)

    # 1b) Determine which rows to exclude from output. Two reasons now:
    #       (a) the manual "Remove" column marked with X (case-insensitive), and
    #       (b) the cost change happened on the same Central calendar day the PO
    #           was created (the _exclude_same_date flag set in normalization).
    #     Excluded rows go to the Excluded tab; Source Data still shows
    #     everything. (PurchaseDetailMatchFound is intentionally ignored.)
    remove_col = next(
        (c for c in merged.columns if str(c).strip().lower() == "remove"),
        None,
    )
    if remove_col is not None:
        remove_mask = merged[remove_col].astype("string").str.strip().str.lower().eq("x").fillna(False)
    else:
        remove_mask = pd.Series(False, index=merged.index)

    if EXCLUDE_SAME_DATE_COL in merged.columns:
        same_date_mask = merged[EXCLUDE_SAME_DATE_COL].fillna(False).astype(bool)
    else:
        same_date_mask = pd.Series(False, index=merged.index)

    excluded_mask = remove_mask | same_date_mask

    # The same-date flag was only needed to build the mask; drop it now so it
    # never appears on Source Data / Excluded sheets.
    merged = merged.drop(columns=[EXCLUDE_SAME_DATE_COL], errors="ignore")

    if excluded_mask.any():
        excluded_raw = merged[excluded_mask].reset_index(drop=True)
        merged_for_pipeline = merged[~excluded_mask].reset_index(drop=True)
        log.info(
            "Excluded %d rows (%d Remove=X, %d same created/adjusted day) covering %d PO #s",
            int(excluded_mask.sum()),
            int(remove_mask.sum()),
            int(same_date_mask.sum()),
            excluded_raw["PO #"].nunique() if len(excluded_raw) else 0,
        )
    else:
        excluded_raw = merged.iloc[0:0].copy()
        merged_for_pipeline = merged

    # 1c) Apply the Cancelled override to the Source Data and Excluded raw
    #     views, so their Total Adjustment column reconciles with the
    #     Combined ledger (which already applies the same override inside
    #     transform()). Doesn't touch the pipeline path — transform() applies
    #     its own override there.
    source_data_view = _apply_cancelled_override_raw(merged)
    excluded_view = _apply_cancelled_override_raw(excluded_raw)

    # 1d) Tag every raw row with its short per-company label (same label used
    #     for the per-company tabs/files), so the per-company download files
    #     can carry that company's own Source Data and Excluded tabs. Mirrors
    #     transform()'s company chain: raw → QBO (master) → canonical casing →
    #     display label → file_label. Hidden helper column, dropped before any
    #     sheet is written.
    _map = get_mapping()
    _canonical_by_lower = {n.lower(): n for n in set(_map.values())}

    def _row_label(raw) -> str:
        if raw is None or (isinstance(raw, float) and pd.isna(raw)):
            return ""
        q = _map.get(str(raw).strip().lower(), raw)
        q = _canonical_by_lower.get(str(q).strip().lower(), q)
        return file_label(display_name(q))

    source_data_view = source_data_view.copy()
    excluded_view = excluded_view.copy()
    source_data_view["_label"] = source_data_view["Company"].map(_row_label)
    excluded_view["_label"] = excluded_view["Company"].map(_row_label)

    # 2) Run the canonical pipeline on the kept rows.
    cleaned, dropped = transform(merged_for_pipeline)

    # 3) Date range from the Adjustment Date column.
    date_range_str = _format_date_range(
        [d for d in cleaned["Adjustment Date"].tolist() if pd.notna(d)]
    )

    # 4) Per-company DataFrames (only ones with data), ordered by display
    #    order so the dict iteration matches the UI grid and tab order.
    grouped_dfs: dict[str, pd.DataFrame] = {}
    for company, grp in cleaned.groupby(cleaned["Company"].map(file_label), dropna=False):
        if pd.isna(company):
            continue
        grouped_dfs[str(company)] = grp.reset_index(drop=True)
    company_dfs: dict[str, pd.DataFrame] = {
        k: grouped_dfs[k] for k in sorted(grouped_dfs.keys(), key=_sort_key)
    }

    # 5) All canonical QBO companies from the master file, translated to
    #    display labels, ordered to match the Purchase Details processor's
    #    tab order (YourTickets last). Used by the UI to render the full grid
    #    and by the combined workbook for tab order.
    all_companies = sorted(
        {file_label(display_name(n)) for n in get_mapping().values()},
        key=_sort_key,
    )

    # 6) Stats block matching the reference's shape.
    #    Note: reference uses "total_cost"; we keep that key name so the UI
    #    template doesn't need editing. The value is sum of Total Adjustment.
    stats: dict[str, dict] = {
        "Combined": {
            "rows": int(len(cleaned)),
            "total_cost": round(float(cleaned["Total Adjustment"].sum()), 2) if len(cleaned) else 0.0,
        }
    }
    for name in all_companies:
        cdf = company_dfs.get(name)
        if cdf is not None and len(cdf) > 0:
            stats[name] = {
                "rows": int(len(cdf)),
                "total_cost": round(float(cdf["Total Adjustment"].sum()), 2),
            }
        else:
            stats[name] = {"rows": 0, "total_cost": 0.0}

    # 7) Build Bills (one row per positive event) and Expenses (debit/credit
    # Output files (combined workbook, per-company expenses files, per-company
    # bills files) are built later by build_filtered_outputs(), once the user
    # has chosen which companies to include. process_files only computes the
    # data + stats needed to show the selection modal.

    # Excluded $ total — same Cancelled-override convention as the Combined
    # ledger, so Source total − Excluded total = Combined total.
    excluded_total = (
        float(pd.to_numeric(excluded_view["Total Adjustment"], errors="coerce").fillna(0).sum())
        if len(excluded_view) else 0.0
    )

    return {
        "date_range": date_range_str,
        "all_companies": all_companies,
        "stats": stats,
        "dropped": dropped,
        "excluded": {
            "po_count": int(excluded_raw["PO #"].nunique()) if len(excluded_raw) else 0,
            "row_count": int(len(excluded_raw)),
            "total_adjustment": round(excluded_total, 2),
        },
        "ignored_companies": out_of_scope_counts,
        # Intermediates for deferred, company-filtered output building. The
        # app pickles these and calls build_filtered_outputs() after the user
        # picks companies in the selection modal.
        "_cleaned": cleaned,
        "_source_view": source_data_view,
        "_excluded_view": excluded_view,
    }


def build_filtered_outputs(
    cleaned: pd.DataFrame,
    source_view: pd.DataFrame,
    excluded_view: pd.DataFrame,
    date_range: str,
    selected_companies: list[str],
) -> dict:
    """Build the output files for the chosen companies only.

    Returns a dict with:
      - "combined"    : the combined workbook bytes (Source Data / Excluded are
                        full; Combined / Summary / Bills / per-company tabs are
                        limited to the selected companies)
      - "companies"   : {label: xlsx bytes} per-company Expenses files (selected
                        companies that have expense rows)
      - "bills_files" : {label: xlsx bytes} per-company PD-format bills files
                        (selected companies that have positive adjustments)

    `selected_companies` are the short sheet/file labels (e.g. "GK", "Y&S").
    """
    selected = set(selected_companies)

    if cleaned.empty:
        cleaned_sel = cleaned
    else:
        cleaned_sel = cleaned[cleaned["Company"].map(file_label).isin(selected)].reset_index(drop=True)

    bills_df, expenses_df = _build_bills_and_expenses(cleaned_sel)
    pd_bills = build_pd_bills(cleaned_sel)

    # Selected labels in display order (drives the combined per-company tabs).
    selected_ordered = sorted(selected, key=_sort_key)

    combined_bytes = _build_combined_workbook(
        source_view, bills_df, expenses_df, selected_ordered, excluded_view,
        pd_bills_df=pd_bills,
    )

    def _company_ledger(name: str) -> pd.DataFrame:
        b = bills_df[bills_df["_display_label"] == name] if "_display_label" in bills_df.columns else bills_df.iloc[0:0]
        e = expenses_df[expenses_df["_display_label"] == name] if "_display_label" in expenses_df.columns else expenses_df.iloc[0:0]
        return _combined_ledger(b, e)

    companies_with_expenses = (
        set(expenses_df["_display_label"]) if "_display_label" in expenses_df.columns else set()
    )
    company_files = {
        name: _build_company_workbook(
            name, expenses_df, source_view, excluded_view, _company_ledger(name)
        )
        for name in selected_ordered
        if name in companies_with_expenses
    }

    bills_files: dict[str, bytes] = {}
    if not pd_bills.empty:
        for label, grp in pd_bills.groupby("_display_label", sort=False):
            if str(label) not in selected:
                continue
            grp_out = grp.drop(columns=["_display_label"])[PD_BILLS_COLUMNS].reset_index(drop=True)
            bills_files[str(label)] = _build_company_file(
                "Bills", grp_out, str(label), source_view, excluded_view,
                _company_ledger(str(label)),
            )
    bills_files = {k: bills_files[k] for k in sorted(bills_files.keys(), key=_sort_key)}

    return {
        "combined": combined_bytes,
        "companies": company_files,
        "bills_files": bills_files,
    }


def convert_to_modified(file_bytes: bytes, filename: str) -> bytes:
    """Zone 1 conversion: reformat one raw PO Cost Changes export into the
    friendly **Modified** review layout — renamed/reordered columns, the
    cost/qty fields split into Start (initial) and End (current), and the five
    internal ID/match columns dropped. Values and seat-level rows are preserved
    exactly (no aggregation, no date conversion).

    The user can review/edit this file and feed it into Zone 2, which reads the
    Modified layout natively. Zone 1 → Zone 2 is equivalent to uploading the
    raw export straight into Zone 2.
    """
    raw = _read_raw(file_bytes, filename)

    # Already a Modified file? Just normalize column order and pass through.
    if _is_modified_format(raw):
        out = pd.DataFrame({c: (raw[c] if c in raw.columns else pd.NA) for c in MODIFIED_COLUMNS})
    else:
        out = pd.DataFrame()
        for raw_col, mod_col in RAW_TO_MODIFIED:
            out[mod_col] = raw[raw_col] if raw_col in raw.columns else pd.NA

    # Cancelled: show the pipeline's Yes/blank convention instead of the raw
    # True/False from the export (all output files use Yes/blank). Non-cancelled
    # rows become truly-empty cells so the column filters cleanly. Round-trips
    # back through Zone 2 unchanged (_to_cancelled maps "Yes"/blank correctly).
    if "Cancelled" in out.columns:
        out["Cancelled"] = [
            "Yes" if _to_cancelled(v) == "Yes" else None
            for v in out["Cancelled"]
        ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Converted"
    _write_modified_sheet(ws, out)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_modified_sheet(ws, df: pd.DataFrame) -> None:
    """Write the Modified review sheet: header band + values exactly as read
    (NaN/NaT → blank). Light styling, frozen header, autosized columns, an
    auto-filter on the header row, and a yellow highlight on the User column for
    rows whose user is a known accounting-team user (see ACCOUNTING_USERS)."""
    from openpyxl.styles import Font, PatternFill, Alignment

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="3F51B5")
    cols = list(df.columns)
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    user_ci = cols.index("User") + 1 if "User" in cols else None

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col in enumerate(cols, 1):
            v = row[col]
            if v is None or (not isinstance(v, (list, tuple)) and pd.isna(v)):
                v = None
            elif isinstance(v, (pd.Timestamp,)):
                v = v.to_pydatetime()
            cell = ws.cell(row=ri, column=ci, value=v)
            # Flag accounting-team users so reviewers can mark Remove = X.
            if ci == user_ci and _is_accounting_user(v):
                cell.fill = ACCOUNTING_FILL

    ws.freeze_panes = "A2"
    for ci, col in enumerate(cols, 1):
        width = max(len(str(col)) + 2, 12)
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(width, 40)

    # Filter dropdowns on the header row (whole used range, or just the header
    # when there are no data rows).
    ws.auto_filter.ref = ws.dimensions

