"""Tests for the v2 PO Cost Changes pipeline (new DB export).

Run with:  pytest

Covers the parts that changed for the new source:
  * _read_one / _normalize_new_export — column mapping, derived Total
    Adjustment (End - Start), UTC → US Central date conversion, IsCancelled
    mapping, manual Remove column passthrough.
  * transform — cancellation override, zero-filter, aggregation.
  * process_files — Remove=X exclusion, PurchaseDetailMatchFound ignored,
    Source/Combined reconciliation — exercised against the real sample files
    when they're available.
"""
import io
import os
import glob

import pandas as pd
import pytest

import processor
import mapping
import teams
import vendors

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
SAMPLE_GLOB = os.path.join(os.path.dirname(__file__), "..", "samples", "PO_Cost_Changes_*.xlsx")


@pytest.fixture(autouse=True)
def _point_at_data(monkeypatch):
    """Point the reference-file loaders at the committed data/ files and clear
    their caches so each test sees a clean load."""
    monkeypatch.setenv("MASTER_MAPPING_PATH", os.path.join(DATA_DIR, "Master_Mapping_List.xlsx"))
    monkeypatch.setenv("TEAMS_PATH", os.path.join(DATA_DIR, "major_league_teams.xlsx"))
    monkeypatch.setenv("OPEN_VENDORS_PATH", os.path.join(DATA_DIR, "Vendors_Open.xlsx"))
    mapping.reset_cache()
    teams.reset_cache()
    vendors.reset_cache()
    yield
    mapping.reset_cache()
    teams.reset_cache()
    vendors.reset_cache()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

NEW_COLUMNS = [
    "POTicketGroupID", "PurchaseOrderID", "Vendor", "Section", "Row",
    "StartSeat", "EndSeat", "IsCancelled", "UpdateUser", "TicketCost",
    "InitialTicketCost", "TicketCostTotal", "InitialTicketCostTotal",
    "Quantity", "InitialQuantity", "AdjustedDateTimeUTC", "PerformerName",
    "SecondaryPerformerName", "EventDate", "EventTime", "CompanyName",
    "ExtPONumber", "VenueName", "AccountEmail", "CreatedDate",
    "PurchaseDetailMatchFound", "MatchedPurchaseOrderID", "MatchedPurchaseDetailID",
]


def _row(**overrides):
    base = {c: None for c in NEW_COLUMNS}
    base.update({
        "POTicketGroupID": 1, "PurchaseOrderID": 1000, "Vendor": "SeatGeek",
        "IsCancelled": False, "UpdateUser": "user1",
        "TicketCostTotal": 100.0, "InitialTicketCostTotal": 60.0,
        "AdjustedDateTimeUTC": "2026-05-30T12:00:00",
        "PerformerName": "Some Show", "CompanyName": "YSA",
        "PurchaseDetailMatchFound": True,
    })
    base.update(overrides)
    return base


def _to_xlsx_bytes(rows, extra_cols=None):
    cols = NEW_COLUMNS + (extra_cols or [])
    df = pd.DataFrame(rows, columns=cols)
    buf = io.BytesIO()
    df.to_excel(buf, sheet_name="Sheet", index=False)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Reader / normalizer
# ---------------------------------------------------------------------------

def test_column_mapping_and_derived_adjustment():
    df = processor._read_one(_to_xlsx_bytes([_row()]), "f.xlsx")
    assert set(["Company", "PO #", "Adjustment Date", "Vendor", "Team/Performer",
                "Ticket Cost Total Start", "Ticket Cost Total End",
                "Total Adjustment", "Cancelled", "User"]).issubset(df.columns)
    r = df.iloc[0]
    assert r["Company"] == "YSA"
    assert r["PO #"] == 1000
    assert r["Ticket Cost Total Start"] == 60.0
    assert r["Ticket Cost Total End"] == 100.0
    # End - Start
    assert r["Total Adjustment"] == 40.0


def test_negative_adjustment_when_cost_drops():
    df = processor._read_one(
        _to_xlsx_bytes([_row(InitialTicketCostTotal=200.0, TicketCostTotal=50.0)]), "f.xlsx"
    )
    assert df.iloc[0]["Total Adjustment"] == -150.0


def test_iscancelled_maps_to_yes_blank():
    df = processor._read_one(
        _to_xlsx_bytes([_row(IsCancelled=True), _row(IsCancelled=False)]), "f.xlsx"
    )
    assert df.iloc[0]["Cancelled"] == "Yes"
    assert df.iloc[1]["Cancelled"] == ""


def test_utc_converted_to_central_can_shift_date():
    # No date in filename → falls back to the UTC → Central conversion.
    # 01:45 UTC on 05-30 is 20:45 CDT on 05-29 → date should be 05-29.
    df = processor._read_one(
        _to_xlsx_bytes([_row(AdjustedDateTimeUTC="2026-05-30T01:45:00")]), "export.xlsx"
    )
    assert df.iloc[0]["Adjustment Date"] == pd.Timestamp("2026-05-29")
    # A midday UTC time stays on the same calendar day.
    df2 = processor._read_one(
        _to_xlsx_bytes([_row(AdjustedDateTimeUTC="2026-05-30T18:00:00")]), "export.xlsx"
    )
    assert df2.iloc[0]["Adjustment Date"] == pd.Timestamp("2026-05-30")


def test_filename_date_overrides_to_that_date():
    # Filename carries a date → every row gets that exact date, regardless of
    # the UTC timestamp (here a midday time that would otherwise be 05-30).
    df = processor._read_one(
        _to_xlsx_bytes([
            _row(AdjustedDateTimeUTC="2026-05-30T18:00:00"),
            _row(AdjustedDateTimeUTC="2026-05-30T02:00:00"),
        ]),
        "PO_Cost_Changes_2026-05-30.xlsx",
    )
    assert list(df["Adjustment Date"].dt.date.astype(str).unique()) == ["2026-05-30"]


def test_filename_date_override_with_duplicate_suffix():
    # A "(1)" duplicate-download suffix doesn't affect the parsed date.
    df = processor._read_one(
        _to_xlsx_bytes([_row()]), "PO_Cost_Changes_2026-06-02 (1).xlsx"
    )
    assert df.iloc[0]["Adjustment Date"] == pd.Timestamp("2026-06-02")


def test_remove_column_passed_through_when_present():
    df = processor._read_one(
        _to_xlsx_bytes([_row()], extra_cols=["Remove"]), "f.xlsx"
    )
    # Column present (value None here) — process_files reads it case-insensitively.
    assert "Remove" in df.columns


def test_missing_required_columns_raises():
    bad = pd.DataFrame({"Foo": [1], "Bar": [2]})
    buf = io.BytesIO()
    bad.to_excel(buf, sheet_name="Sheet", index=False)
    with pytest.raises(ValueError):
        processor._read_one(buf.getvalue(), "bad.xlsx")


# ---------------------------------------------------------------------------
# transform
# ---------------------------------------------------------------------------

def _norm(rows):
    return processor._read_one(_to_xlsx_bytes(rows), "f.xlsx")


def test_cancelled_override_reverses_total_end():
    # Cancelled row: adjustment becomes -(Total End), regardless of derived value.
    df = _norm([_row(IsCancelled=True, InitialTicketCostTotal=100.0, TicketCostTotal=30.0)])
    cleaned, _ = processor.transform(df)
    assert cleaned.iloc[0]["Total Adjustment"] == -30.0


def test_zero_adjustment_rows_dropped():
    df = _norm([_row(InitialTicketCostTotal=100.0, TicketCostTotal=100.0)])
    cleaned, _ = processor.transform(df)
    assert cleaned.empty


def test_same_key_rows_aggregate():
    rows = [
        _row(PurchaseOrderID=1, InitialTicketCostTotal=0, TicketCostTotal=50,
             Vendor="SeatGeek", PerformerName="Show A"),
        _row(PurchaseOrderID=2, InitialTicketCostTotal=0, TicketCostTotal=30,
             Vendor="SeatGeek", PerformerName="Show A"),
    ]
    df = _norm(rows)
    cleaned, _ = processor.transform(df)
    # Same (Company, Date, Vendor, Team/Performer) → one aggregated row of 80.
    assert len(cleaned) == 1
    assert cleaned.iloc[0]["Total Adjustment"] == 80.0


def test_unmapped_company_dropped():
    df = _norm([_row(CompanyName="Totally Unknown Co", TicketCostTotal=100, InitialTicketCostTotal=0)])
    cleaned, dropped = processor.transform(df)
    assert cleaned.empty
    assert dropped["total_dropped_rows"] == 1


def test_memo_format():
    df = _norm([_row(CompanyName="YSA", Vendor="SeatGeek",
                     PerformerName="Olivia Rodrigo",
                     AccountEmail="walter@outlook.com", ExtPONumber="ABC-123",
                     InitialTicketCostTotal=0, TicketCostTotal=100)])
    cleaned, _ = processor.transform(df)
    bills, _expenses = processor._build_bills_and_expenses(cleaned)
    assert bills.iloc[0]["Memo"] == "Olivia Rodrigo / walter@outlook.com / ABC-123 / Cost Changes (YSA)"
    # Description mirrors Memo.
    assert bills.iloc[0]["Description"] == bills.iloc[0]["Memo"]


def test_memo_blank_extpo_omits_segment():
    df = _norm([_row(CompanyName="YSA", PerformerName="Hamilton",
                     AccountEmail="x@y.com", ExtPONumber=None,
                     InitialTicketCostTotal=0, TicketCostTotal=80)])
    cleaned, _ = processor.transform(df)
    bills, _ = processor._build_bills_and_expenses(cleaned)
    # No ext PO → that segment is omitted entirely (no empty " /  / ").
    assert bills.iloc[0]["Memo"] == "Hamilton / x@y.com / Cost Changes (YSA)"


def test_memo_blank_email_and_extpo_omits_both():
    df = _norm([_row(CompanyName="YSA", PerformerName="Chicago Cubs",
                     AccountEmail=None, ExtPONumber=None,
                     InitialTicketCostTotal=0, TicketCostTotal=80)])
    cleaned, _ = processor.transform(df)
    bills, _ = processor._build_bills_and_expenses(cleaned)
    assert bills.iloc[0]["Memo"] == "Chicago Cubs / Cost Changes (YSA)"


def test_aggregation_splits_by_account_email():
    rows = [
        _row(PurchaseOrderID=1, InitialTicketCostTotal=0, TicketCostTotal=50,
             AccountEmail="a@x.com", PerformerName="Show"),
        _row(PurchaseOrderID=1, InitialTicketCostTotal=0, TicketCostTotal=30,
             AccountEmail="b@x.com", PerformerName="Show"),
    ]
    cleaned, _ = processor.transform(_norm(rows))
    # Different emails → two separate output rows (not summed).
    assert len(cleaned) == 2


def test_aggregation_splits_by_extpo():
    rows = [
        _row(PurchaseOrderID=1, InitialTicketCostTotal=0, TicketCostTotal=50,
             AccountEmail="a@x.com", ExtPONumber="PO-1", PerformerName="Show"),
        _row(PurchaseOrderID=1, InitialTicketCostTotal=0, TicketCostTotal=30,
             AccountEmail="a@x.com", ExtPONumber="PO-2", PerformerName="Show"),
    ]
    cleaned, _ = processor.transform(_norm(rows))
    assert len(cleaned) == 2


def test_same_email_and_extpo_still_aggregate():
    rows = [
        _row(PurchaseOrderID=1, InitialTicketCostTotal=0, TicketCostTotal=50,
             AccountEmail="a@x.com", ExtPONumber="PO-1", PerformerName="Show"),
        _row(PurchaseOrderID=2, InitialTicketCostTotal=0, TicketCostTotal=30,
             AccountEmail="a@x.com", ExtPONumber="PO-1", PerformerName="Show"),
    ]
    cleaned, _ = processor.transform(_norm(rows))
    # Same full key (incl. email + extPO) → summed into one row of 80.
    assert len(cleaned) == 1
    assert cleaned.iloc[0]["Total Adjustment"] == 80.0


# ---------------------------------------------------------------------------
# Vendor pipeline (ported from Purchase Details app)
# ---------------------------------------------------------------------------

import vendor_rules


def _final_vendor(rows):
    """Run transform on rows and return the set of resulting Vendor values."""
    cleaned, _ = processor.transform(_norm(rows))
    return list(cleaned["Vendor"])


def test_ticketmaster_am_to_team():
    v = _final_vendor([_row(Vendor="Ticketmaster AM", PerformerName="Los Angeles Dodgers",
                            VenueName="Dodger Stadium",
                            InitialTicketCostTotal=0, TicketCostTotal=100)])
    assert v == ["Los Angeles Dodgers"]


def test_ticketmaster_am_to_venue_when_not_team():
    # Performer not a major-league team → falls back to venue (then title-cased).
    v = _final_vendor([_row(Vendor="Ballpark", PerformerName="Gracie Abrams",
                            VenueName="Madison Square Garden",
                            InitialTicketCostTotal=0, TicketCostTotal=100)])
    assert v == ["Madison Square Garden"]


def test_axs_substring_replacement_to_veritix():
    v = _final_vendor([_row(Vendor="AXS", PerformerName="Some Act", VenueName="Some Venue",
                            InitialTicketCostTotal=0, TicketCostTotal=100)])
    assert v == ["Veritix"]


def test_concert_extras_at_msg_becomes_madison_square_garden():
    # At any MSG Entertainment venue (MSG + its parking lots, Radio City Music
    # Hall, Beacon Theatre), Concert Extras -> "Madison Square Garden",
    # overriding the general Concert Extras -> Live Nation Extras rule.
    for ven in ("Madison Square Garden", "Madison Square Garden Parking Lots",
                "Radio City Music Hall", "Beacon Theatre - New York"):
        v = _final_vendor([_row(Vendor="Concert Extras", PerformerName="Some Act",
                                VenueName=ven, InitialTicketCostTotal=0, TicketCostTotal=100)])
        assert v == ["Madison Square Garden"], ven


def test_concert_extras_elsewhere_still_live_nation_extras():
    v = _final_vendor([_row(Vendor="Concert Extras", PerformerName="Some Act",
                            VenueName="Some Other Arena", InitialTicketCostTotal=0,
                            TicketCostTotal=100)])
    assert v == ["Live Nation Extras"]


def test_tickets_com_team_vs_venue():
    rows = [
        _row(PurchaseOrderID=1, Vendor="Tickets.com", PerformerName="New York Yankees",
             VenueName="Yankee Stadium", AccountEmail="a@x.com",
             InitialTicketCostTotal=0, TicketCostTotal=50),
        _row(PurchaseOrderID=2, Vendor="Tickets.com", PerformerName="Not A Team",
             VenueName="Some Theater", AccountEmail="b@x.com",
             InitialTicketCostTotal=0, TicketCostTotal=50),
    ]
    vendors = set(_final_vendor(rows))
    assert "New York Yankees" in vendors
    assert "Some Theater" in vendors


def test_ysa_live_nation_becomes_concert_seasons_then_venue_map():
    # YSA + Live Nation → Concert Seasons → CONCERT_SEASONS_MAP[venue].
    v = _final_vendor([_row(CompanyName="YSA", Vendor="Live Nation",
                            PerformerName="Some Act", VenueName="Concord Pavilion",
                            InitialTicketCostTotal=0, TicketCostTotal=100)])
    assert v == ["Live Nation Concord Pavilion"]


def test_clean_ext_po_blanks_uuid_and_long_numeric():
    df = pd.DataFrame({
        "Vendor": ["SeatGeek", "SeatGeek", "SeatGeek"],
        "ExtPONumber": [
            "12345678-1234-1234-1234-123456789abc",  # uuid → blank
            "1234567890123456789",                    # 19-digit → blank
            "ABC-123",                                # keep
        ],
    })
    out = vendor_rules.clean_ext_po(df.copy())
    assert out["ExtPONumber"].tolist() == ["", "", "ABC-123"]


def test_clean_ext_po_concert_seasons_always_tmam_only_15plus():
    df = pd.DataFrame({
        "Vendor": [
            "Concert Seasons",   # always blanked
            "Ticketmaster AM",   # short alphanumeric → kept
            "Ticketmaster AM",   # 15-digit numeric → blanked
            "Ticketmaster AM",   # 14-digit numeric → kept (below threshold)
            "SeatGeek",          # untouched
        ],
        "ExtPONumber": [
            "KEEP-1",
            "ABC-123",
            "123456789012345",   # 15 digits
            "12345678901234",    # 14 digits
            "KEEP-5",
        ],
    })
    out = vendor_rules.clean_ext_po(df.copy())
    assert out["ExtPONumber"].tolist() == ["", "ABC-123", "", "12345678901234", "KEEP-5"]


# ---------------------------------------------------------------------------
# PD-format per-company bills files
# ---------------------------------------------------------------------------

def test_pd_bills_columns_and_values():
    rows = [_row(CompanyName="YSA", Vendor="SeatGeek", PerformerName="Olivia Rodrigo",
                 AccountEmail="a@b.com", ExtPONumber="PO-9",
                 InitialTicketCostTotal=0, TicketCostTotal=100)]
    cleaned, _ = processor.transform(_norm(rows))
    pdb = processor.build_pd_bills(cleaned)
    # Column order matches the PD layout (plus the hidden split helper).
    assert list(pdb.columns) == processor.PD_BILLS_COLUMNS + ["_display_label"]
    r = pdb.iloc[0]
    assert r["Company"] == "YSA"            # raw/original company
    assert r["Account"] == "Inventory Asset"
    assert r["Memo2"] == "Olivia Rodrigo / a@b.com / PO-9"          # no company
    assert r["Team/Performer"] == "Olivia Rodrigo / a@b.com / PO-9 / Cost Changes (YSA)"
    assert r["Memo"] == r["Team/Performer"]
    assert r["Total Cost"] == 100
    assert r["Seasons"] == ""
    assert 10000000 <= int(r["Bill No."]) <= 99999999


def test_pd_bills_excludes_negative_adjustments():
    rows = [
        _row(PurchaseOrderID=1, PerformerName="Pos", AccountEmail="a@b.com",
             InitialTicketCostTotal=0, TicketCostTotal=100),   # +100 → bill
        _row(PurchaseOrderID=2, PerformerName="Neg", AccountEmail="c@d.com",
             InitialTicketCostTotal=100, TicketCostTotal=20),  # -80 → expense, excluded
    ]
    cleaned, _ = processor.transform(_norm(rows))
    pdb = processor.build_pd_bills(cleaned)
    assert len(pdb) == 1
    assert pdb.iloc[0]["Total Cost"] == 100


def test_pd_bills_per_company_files_tie_to_bills_tab():
    rows = [
        _row(PurchaseOrderID=1, CompanyName="YSA", PerformerName="A", AccountEmail="a@b.com",
             InitialTicketCostTotal=0, TicketCostTotal=100),
        _row(PurchaseOrderID=2, CompanyName="YS Katz", PerformerName="B", AccountEmail="b@b.com",
             InitialTicketCostTotal=0, TicketCostTotal=60),
    ]
    res = processor.process_files([(_to_xlsx_bytes(rows), "f.xlsx")])
    out = processor.build_filtered_outputs(
        res["_cleaned"], res["_source_view"], res["_excluded_view"],
        res["date_range"], res["all_companies"],
    )
    # One bills file per company with positive adjustments.
    assert set(out["bills_files"].keys()) == {"Asher", "Katz"}


def test_build_filtered_outputs_respects_selection():
    rows = [
        _row(PurchaseOrderID=1, CompanyName="YSA", PerformerName="A", AccountEmail="a@b.com",
             InitialTicketCostTotal=0, TicketCostTotal=100),
        _row(PurchaseOrderID=2, CompanyName="YS Katz", PerformerName="B", AccountEmail="b@b.com",
             InitialTicketCostTotal=0, TicketCostTotal=60),
    ]
    res = processor.process_files([(_to_xlsx_bytes(rows), "f.xlsx")])
    # Select only YSA → only YSA's bills file is produced.
    out = processor.build_filtered_outputs(
        res["_cleaned"], res["_source_view"], res["_excluded_view"],
        res["date_range"], ["Asher"],
    )
    assert set(out["bills_files"].keys()) == {"Asher"}
    assert out["combined"]


def test_memo_includes_single_created_date():
    rows = [_row(CompanyName="YSA", PerformerName="Hamilton", AccountEmail="a@b.com",
                 ExtPONumber="PO-1",
                 AdjustedDateTimeUTC="2026-05-30T18:00:00",
                 CreatedDate="2026-03-09T12:00:00",
                 InitialTicketCostTotal=0, TicketCostTotal=100)]
    cleaned, _ = processor.transform(_norm(rows))
    bills, _ = processor._build_bills_and_expenses(cleaned)
    assert bills.iloc[0]["Memo"] == (
        "Hamilton / a@b.com / PO-1 / Cost Changes (YSA) (PO created date 03/09/2026)"
    )


def test_memo_lists_multiple_created_dates():
    # Two rows that aggregate into one (same keys) but different created dates.
    rows = [
        _row(PurchaseOrderID=1, CompanyName="YSA", PerformerName="Hamilton",
             AccountEmail="a@b.com", ExtPONumber="PO-1",
             AdjustedDateTimeUTC="2026-05-30T18:00:00",
             CreatedDate="2026-03-09T12:00:00",
             InitialTicketCostTotal=0, TicketCostTotal=50),
        _row(PurchaseOrderID=2, CompanyName="YSA", PerformerName="Hamilton",
             AccountEmail="a@b.com", ExtPONumber="PO-1",
             AdjustedDateTimeUTC="2026-05-30T18:00:00",
             CreatedDate="2026-03-23T12:00:00",
             InitialTicketCostTotal=0, TicketCostTotal=30),
    ]
    cleaned, _ = processor.transform(_norm(rows))
    assert len(cleaned) == 1  # aggregated
    bills, _ = processor._build_bills_and_expenses(cleaned)
    memo = bills.iloc[0]["Memo"]
    assert "(PO created date 03/09/2026, 03/23/2026)" in memo


def test_memo2_includes_created_date_no_company():
    rows = [_row(CompanyName="YSA", PerformerName="Hamilton", AccountEmail="a@b.com",
                 ExtPONumber=None,
                 AdjustedDateTimeUTC="2026-05-30T18:00:00",
                 CreatedDate="2026-03-09T12:00:00",
                 InitialTicketCostTotal=0, TicketCostTotal=100)]
    cleaned, _ = processor.transform(_norm(rows))
    pdb = processor.build_pd_bills(cleaned)
    # Memo2: no company, no "Cost Changes", but created date present.
    assert pdb.iloc[0]["Memo2"] == "Hamilton / a@b.com (PO created date 03/09/2026)"


def test_company_value_rename_and_labels():
    # The four PD-renamed companies: Company-column value vs short sheet label.
    assert processor.company_value("GK LLC") == "YSKG"
    assert processor.company_value("Jacks YS") == "Chase (Jacks)"
    assert processor.company_value("YSW") == "YSW (Waxler)"
    assert processor.company_value("The Ticket Guy") == "Ticket Guy"
    assert processor.company_value("The Ticket Guy VIP") == "Ticket Guy"
    # Everyone else keeps their raw company name.
    assert processor.company_value("YS-Seatgeek2") == "YS-Seatgeek2"
    # Short sheet/file labels.
    assert processor.file_label("YSKG") == "GK"
    assert processor.file_label("Chase (Jacks)") == "Chase"
    assert processor.file_label("YSW (Waxler)") == "Waxler"
    assert processor.file_label("Ticket Guy") == "Ticket Guy"
    assert processor.file_label("Y&S") == "Y&S"


def test_company_files_are_expenses_only():
    rows = [
        _row(PurchaseOrderID=1, CompanyName="GK LLC", PerformerName="A",
             AccountEmail="a@b.com", InitialTicketCostTotal=100, TicketCostTotal=20),  # -80 expense
    ]
    res = processor.process_files([(_to_xlsx_bytes(rows), "f.xlsx")])
    out = processor.build_filtered_outputs(
        res["_cleaned"], res["_source_view"], res["_excluded_view"],
        res["date_range"], res["all_companies"],
    )
    # File/label is the short "GK"; tab 1 Expenses, tabs 2/3 Source Data + Excluded.
    assert "GK" in out["companies"]
    import openpyxl, io as _io
    wb = openpyxl.load_workbook(_io.BytesIO(out["companies"]["GK"]))
    assert wb.sheetnames == ["Expenses", "Summary", "Source Data", "Excluded"]
    ws = wb["Expenses"]
    h = [c.value for c in ws[1]]
    ci = h.index("Company")
    vals = {r[ci] for r in ws.iter_rows(min_row=2, values_only=True) if r[ci]}
    assert vals == {"YSKG"}   # Company-column value is the renamed value
    # Source Data tab carries this company's input row(s).
    sd = wb["Source Data"]
    sh = [c.value for c in sd[1]]
    sci = sh.index("Company")
    src_vals = {r[sci] for r in sd.iter_rows(min_row=2, values_only=True) if r[sci]}
    assert src_vals == {"GK LLC"}   # Source Data keeps the raw company name


def test_company_files_have_company_scoped_tabs():
    # Two companies, one with a same-day exclusion. Each company's file should
    # only show its OWN rows on Source Data / Excluded.
    rows = [
        _row(PurchaseOrderID=1, CompanyName="GK LLC", PerformerName="A",
             AccountEmail="a@b.com", InitialTicketCostTotal=100, TicketCostTotal=20),
        _row(PurchaseOrderID=2, CompanyName="YSA", PerformerName="B",
             AccountEmail="b@b.com", InitialTicketCostTotal=50, TicketCostTotal=10),
    ]
    res = processor.process_files([(_to_xlsx_bytes(rows), "f.xlsx")])
    out = processor.build_filtered_outputs(
        res["_cleaned"], res["_source_view"], res["_excluded_view"],
        res["date_range"], res["all_companies"],
    )
    import openpyxl, io as _io
    for label, raw_company in [("GK", "GK LLC"), ("Asher", "YSA")]:
        wb = openpyxl.load_workbook(_io.BytesIO(out["companies"][label]))
        assert wb.sheetnames == ["Expenses", "Summary", "Source Data", "Excluded"]
        sd = wb["Source Data"]
        sci = [c.value for c in sd[1]].index("Company")
        src_vals = {r[sci] for r in sd.iter_rows(min_row=2, values_only=True) if r[sci]}
        assert src_vals == {raw_company}


def test_offset_category_inverted_rule():
    # Vendors ON the third-tab TC list → "<vendor> (TC)"; everything else →
    # "Due from Vendors - Open".
    tc = vendors.get_tc_vendors()
    assert len(tc) > 0
    assert "anaheim ducks" in tc                       # known third-tab entry
    assert vendors.offset_category("Anaheim Ducks", tc) == "Anaheim Ducks (TC)"
    assert vendors.offset_category("anaheim ducks", tc) == "anaheim ducks (TC)"
    assert vendors.offset_category("Nope Not Listed", tc) == "Due from Vendors - Open"


def test_expense_offset_pair_uses_tc_list():
    # A negative adjustment produces a two-line expense pair: Inventory Asset +
    # the offset category, which depends on whether the final vendor is on TC.
    rows = [_row(CompanyName="YSA", Vendor="SeatGeek", PerformerName="Some Act",
                 AccountEmail="a@b.com",
                 InitialTicketCostTotal=100, TicketCostTotal=20)]   # -80
    cleaned, _ = processor.transform(_norm(rows))
    final_vendor = cleaned.iloc[0]["Vendor"]

    # On the TC list → "<vendor> (TC)"
    _, exp_on = processor._build_bills_and_expenses(cleaned, tc_vendors={final_vendor.lower()})
    cats_on = exp_on["Category"].tolist()
    assert "Inventory Asset" in cats_on
    assert f"{final_vendor} (TC)" in cats_on
    assert "Due from Vendors - Open" not in cats_on

    # Not on the TC list → "Due from Vendors - Open"
    _, exp_off = processor._build_bills_and_expenses(cleaned, tc_vendors=set())
    cats_off = exp_off["Category"].tolist()
    assert "Inventory Asset" in cats_off
    assert "Due from Vendors - Open" in cats_off
    assert f"{final_vendor} (TC)" not in cats_off
    # TM AM with a 15-digit order number and a team performer. Cleaning runs
    # first on the RAW vendor "Ticketmaster AM", so the 15+ digit order is
    # blanked even though the vendor then resolves to the team name.
    rows = [_row(CompanyName="YSA", Vendor="Ticketmaster AM",
                 PerformerName="New York Yankees", VenueName="Yankee Stadium",
                 AccountEmail="a@x.com", ExtPONumber="123456789012345",
                 InitialTicketCostTotal=0, TicketCostTotal=100)]
    cleaned, _ = processor.transform(_norm(rows))
    assert cleaned.iloc[0]["Vendor"] == "New York Yankees"
    assert cleaned.iloc[0]["ExtPONumber"] == ""
    bills, _ = processor._build_bills_and_expenses(cleaned)
    assert bills.iloc[0]["Memo"] == "New York Yankees / a@x.com / Cost Changes (YSA)"


# ---------------------------------------------------------------------------
# process_files
# ---------------------------------------------------------------------------

def test_remove_x_excludes_rows():
    rows = [
        _row(PurchaseOrderID=1, TicketCostTotal=100, InitialTicketCostTotal=0),
        _row(PurchaseOrderID=2, TicketCostTotal=200, InitialTicketCostTotal=0),
    ]
    content = _to_xlsx_bytes(
        [{**r, "Remove": ("X" if r["PurchaseOrderID"] == 2 else "")} for r in rows],
        extra_cols=["Remove"],
    )
    res = processor.process_files([(content, "f.xlsx")])
    assert res["excluded"]["row_count"] == 1
    assert res["excluded"]["po_count"] == 1


def test_files_missing_remove_column_flags_only_files_without_it():
    rows = [_row(PurchaseOrderID=1, TicketCostTotal=100, InitialTicketCostTotal=0)]
    with_remove = _to_xlsx_bytes([{**rows[0], "Remove": ""}], extra_cols=["Remove"])
    without_remove = _to_xlsx_bytes(rows)
    missing = processor.files_missing_remove_column([
        (with_remove, "has_it.xlsx"),
        (without_remove, "missing.xlsx"),
    ])
    assert missing == ["missing.xlsx"]


def test_files_missing_remove_column_empty_when_all_present():
    rows = [_row(PurchaseOrderID=1, TicketCostTotal=100, InitialTicketCostTotal=0)]
    a = _to_xlsx_bytes([{**rows[0], "Remove": ""}], extra_cols=["Remove"])
    b = _to_xlsx_bytes([{**rows[0], "Remove": "X"}], extra_cols=["Remove"])
    assert processor.files_missing_remove_column([(a, "a.xlsx"), (b, "b.xlsx")]) == []


def test_files_missing_remove_column_detects_in_converted_file():
    # A Zone 1 "Converted" workbook with a hand-added Remove column is accepted.
    rows = [_row(PurchaseOrderID=1, TicketCostTotal=100, InitialTicketCostTotal=0)]
    converted = processor.convert_to_modified(_to_xlsx_bytes(rows), "raw.xlsx")
    # Without Remove -> flagged
    assert processor.files_missing_remove_column([(converted, "conv.xlsx")]) == ["conv.xlsx"]
    # Add a Remove column to the converted sheet -> accepted
    import openpyxl, io as _io
    wb = openpyxl.load_workbook(_io.BytesIO(converted))
    ws = wb["Converted"]
    ws.cell(row=1, column=ws.max_column + 1, value="Remove")
    buf = _io.BytesIO(); wb.save(buf)
    assert processor.files_missing_remove_column([(buf.getvalue(), "conv.xlsx")]) == []


def test_purchase_detail_match_is_ignored():
    # All rows have match=True; none should be excluded (we ignore the flag).
    # Distinct performers so they don't collapse in the final aggregation.
    rows = [_row(PurchaseOrderID=i, TicketCostTotal=100, InitialTicketCostTotal=0,
                 PerformerName=f"Show {i}", PurchaseDetailMatchFound=True)
            for i in range(1, 4)]
    res = processor.process_files([(_to_xlsx_bytes(rows), "f.xlsx")])
    assert res["excluded"]["row_count"] == 0
    assert res["stats"]["Combined"]["rows"] == 3


def test_same_created_and_adjusted_day_excluded():
    # Created and adjusted on the same Central day → excluded.
    rows = [
        _row(PurchaseOrderID=1, PerformerName="Same Day",
             AdjustedDateTimeUTC="2026-05-30T18:00:00",   # 13:00 CDT on 05-30
             CreatedDate="2026-05-30T15:00:00",            # 10:00 CDT on 05-30
             TicketCostTotal=100, InitialTicketCostTotal=0),
        _row(PurchaseOrderID=2, PerformerName="Diff Day",
             AdjustedDateTimeUTC="2026-05-30T18:00:00",
             CreatedDate="2026-05-28T15:00:00",            # different day → kept
             TicketCostTotal=100, InitialTicketCostTotal=0),
    ]
    res = processor.process_files([(_to_xlsx_bytes(rows), "f.xlsx")])
    assert res["excluded"]["row_count"] == 1
    assert res["stats"]["Combined"]["rows"] == 1


def test_same_day_comparison_uses_central_not_utc():
    # Adjusted 2026-05-30T02:00 UTC = 05-29 21:00 CDT. Created 2026-05-29T20:00
    # UTC = 05-29 15:00 CDT. Same UTC date? No (05-30 vs 05-29). Same Central
    # date? Yes (both 05-29) → excluded under the Central rule.
    rows = [_row(PurchaseOrderID=1, PerformerName="X",
                 AdjustedDateTimeUTC="2026-05-30T02:00:00",
                 CreatedDate="2026-05-29T20:00:00",
                 TicketCostTotal=100, InitialTicketCostTotal=0)]
    res = processor.process_files([(_to_xlsx_bytes(rows), "f.xlsx")])
    assert res["excluded"]["row_count"] == 1


def test_blank_created_date_is_not_a_same_day_match():
    rows = [_row(PurchaseOrderID=1, PerformerName="X",
                 AdjustedDateTimeUTC="2026-05-30T18:00:00",
                 CreatedDate=None,
                 TicketCostTotal=100, InitialTicketCostTotal=0)]
    res = processor.process_files([(_to_xlsx_bytes(rows), "f.xlsx")])
    assert res["excluded"]["row_count"] == 0
    assert res["stats"]["Combined"]["rows"] == 1


# ---------------------------------------------------------------------------
# Real sample files (skipped if not present)
# ---------------------------------------------------------------------------

@pytest.mark.skipif(not glob.glob(SAMPLE_GLOB), reason="no sample files committed")
def test_real_samples_reconcile():
    files = [(open(f, "rb").read(), os.path.basename(f)) for f in sorted(glob.glob(SAMPLE_GLOB))]
    res = processor.process_files(files)
    out = processor.build_filtered_outputs(
        res["_cleaned"], res["_source_view"], res["_excluded_view"],
        res["date_range"], res["all_companies"],
    )
    assert out["combined"]
    assert res["stats"]["Combined"]["rows"] > 0


# ---------------------------------------------------------------------------
# Live Nation (concert-season) collapse
# ---------------------------------------------------------------------------

def test_live_nation_collapses_to_various():
    rows = [
        _row(PurchaseOrderID=1, CompanyName="YSA", Vendor="Concert Seasons",
             VenueName="Jiffy Lube Live", PerformerName="Band A",
             AccountEmail="a@x.com", ExtPONumber="ORD111",
             InitialTicketCostTotal=0, TicketCostTotal=100),
        _row(PurchaseOrderID=2, CompanyName="YSA", Vendor="Concert Seasons",
             VenueName="Jiffy Lube Live", PerformerName="Band B",
             AccountEmail="b@x.com", ExtPONumber="ORD222",
             InitialTicketCostTotal=0, TicketCostTotal=150),
    ]
    res = processor.process_files([(_to_xlsx_bytes(rows), "PO_Cost_Changes_2026-06-02.xlsx")])
    cl = res["_cleaned"]
    ln = cl[cl["Vendor"].str.contains("Live Nation")]
    assert len(ln) == 1
    r = ln.iloc[0]
    assert r["Team/Performer"] == "Various / Various"
    assert r["AccountEmail"] == "" and r["ExtPONumber"] == ""
    assert r["Total Adjustment"] == 250


def test_non_live_nation_keeps_detail():
    rows = [_row(PurchaseOrderID=1, CompanyName="YSA", Vendor="SeatGeek",
                 PerformerName="Knicks", AccountEmail="c@x.com", ExtPONumber="ORD333",
                 InitialTicketCostTotal=0, TicketCostTotal=50)]
    res = processor.process_files([(_to_xlsx_bytes(rows), "f.xlsx")])
    r = res["_cleaned"].iloc[0]
    assert r["Team/Performer"] == "Knicks"
    assert r["AccountEmail"] == "c@x.com"


# ---------------------------------------------------------------------------
# Re-uploading an app output file (edited Source Data tab) + Zone 1 convert
# ---------------------------------------------------------------------------

def _combined_bytes(rows):
    res = processor.process_files([(_to_xlsx_bytes(rows), "f.xlsx")])
    out = processor.build_filtered_outputs(
        res["_cleaned"], res["_source_view"], res["_excluded_view"],
        res["date_range"], res["all_companies"],
    )
    return res, out["combined"]


def test_reupload_output_file_round_trips():
    rows = [
        _row(PurchaseOrderID=1, CompanyName="GK LLC", PerformerName="A",
             AccountEmail="a@b.com", InitialTicketCostTotal=100, TicketCostTotal=20),
        _row(PurchaseOrderID=2, CompanyName="YSA", PerformerName="B",
             AccountEmail="b@b.com", InitialTicketCostTotal=0, TicketCostTotal=80),
    ]
    res1, combined = _combined_bytes(rows)
    res2 = processor.process_files([(combined, "PO Cost Changes - Combined - June 2nd 2026.xlsx")])
    assert res2["date_range"] == res1["date_range"]
    assert res2["stats"]["Combined"] == res1["stats"]["Combined"]
    assert res2["excluded"]["row_count"] == res1["excluded"]["row_count"]


def test_reupload_detected_as_source_data():
    rows = [_row(PurchaseOrderID=1, CompanyName="GK LLC", PerformerName="A",
                 AccountEmail="a@b.com", InitialTicketCostTotal=100, TicketCostTotal=20)]
    _, combined = _combined_bytes(rows)
    df = processor._read_one(combined, "whatever - Combined - June.xlsx")
    assert {"Company", "PO #", "Vendor", "Ticket Cost Total Start"} <= set(df.columns)


def test_reupload_honors_hand_added_remove_flag():
    import openpyxl, io as _io
    rows = [
        _row(PurchaseOrderID=1, CompanyName="GK LLC", PerformerName="A",
             AccountEmail="a@b.com", InitialTicketCostTotal=100, TicketCostTotal=20),
        _row(PurchaseOrderID=2, CompanyName="GK LLC", PerformerName="B",
             AccountEmail="b@b.com", InitialTicketCostTotal=50, TicketCostTotal=10),
    ]
    res1, combined = _combined_bytes(rows)
    assert res1["excluded"]["row_count"] == 0
    wb = openpyxl.load_workbook(_io.BytesIO(combined))
    ws = wb["Source Data"]
    col = ws.max_column + 1
    ws.cell(row=1, column=col, value="Remove")
    ws.cell(row=2, column=col, value="X")
    buf = _io.BytesIO(); wb.save(buf)
    res2 = processor.process_files([(buf.getvalue(), "edited - Combined - June.xlsx")])
    assert res2["excluded"]["row_count"] == 1


def test_zone1_convert_to_modified_round_trips():
    # Zone 1 reformats raw -> Modified (rename/reorder, values unchanged, no
    # aggregation). Feeding the Modified file into Zone 2 must equal feeding the
    # raw export into Zone 2, including same-day exclusions.
    rows = [
        _row(PurchaseOrderID=1, CompanyName="YSA", PerformerName="A",
             AccountEmail="a@b.com", InitialTicketCostTotal=0, TicketCostTotal=100,
             AdjustedDateTimeUTC="2026-05-30T18:00:00", CreatedDate="2026-05-30T18:00:00"),
        _row(PurchaseOrderID=2, CompanyName="YSA", PerformerName="B",
             AccountEmail="b@b.com", InitialTicketCostTotal=0, TicketCostTotal=80,
             AdjustedDateTimeUTC="2026-05-30T18:00:00", CreatedDate="2026-05-20T18:00:00"),
    ]
    raw_bytes = _to_xlsx_bytes(rows)
    modified = processor.convert_to_modified(raw_bytes, "PO_Cost_Changes_2026-05-30.xlsx")

    import openpyxl, io as _io
    wb = openpyxl.load_workbook(_io.BytesIO(modified))
    assert wb.sheetnames == ["Converted"]
    ws = wb["Converted"]
    header = [c.value for c in ws[1]]
    assert header == processor.MODIFIED_COLUMNS          # exact layout
    assert "Team/Perfomer" in header                     # template spelling kept
    assert ws.max_row - 1 == len(rows)                   # seat-level rows preserved

    raw = processor.process_files([(raw_bytes, "PO_Cost_Changes_2026-05-30.xlsx")])
    mod = processor.process_files([(modified, "x (modified).xlsx")])
    assert raw["stats"]["Combined"] == mod["stats"]["Combined"]
    assert raw["excluded"]["row_count"] == mod["excluded"]["row_count"]
    assert raw["date_range"] == mod["date_range"]
    # Row 1 (same created+adjusted day) excluded; row 2 kept.
    assert raw["excluded"]["row_count"] == 1


def test_zone1_modified_is_pure_reformat_no_aggregation():
    # Two seat lines on the same PO must stay as two rows (no aggregation).
    rows = [
        _row(PurchaseOrderID=9, CompanyName="YSA", PerformerName="A",
             Section="100", Row="A", StartSeat=1, EndSeat=2,
             InitialTicketCostTotal=50, TicketCostTotal=20),
        _row(PurchaseOrderID=9, CompanyName="YSA", PerformerName="A",
             Section="100", Row="B", StartSeat=1, EndSeat=2,
             InitialTicketCostTotal=60, TicketCostTotal=30),
    ]
    modified = processor.convert_to_modified(_to_xlsx_bytes(rows), "f.xlsx")
    import openpyxl, io as _io
    ws = openpyxl.load_workbook(_io.BytesIO(modified))["Converted"]
    assert ws.max_row - 1 == 2
    hdr = [c.value for c in ws[1]]
    # Start = initial, End = current
    start = ws.cell(row=2, column=hdr.index("Total Ticket Start") + 1).value
    end = ws.cell(row=2, column=hdr.index("Total Ticket End") + 1).value
    assert start == 50 and end == 20


# ---------------------------------------------------------------------------
# Zone 1 Converted: accounting-user highlight, header filters, Cancelled
# rendered as Yes/blank (not True/False).
# ---------------------------------------------------------------------------

def _converted_ws(rows, filename="PO_Cost_Changes_2026-05-30.xlsx"):
    import openpyxl
    b = processor.convert_to_modified(_to_xlsx_bytes(rows), filename)
    return openpyxl.load_workbook(io.BytesIO(b))["Converted"]


def _fill_rgb(cell):
    return cell.fill.fgColor.rgb if cell.fill and cell.fill.patternType else None


def test_zone1_highlights_accounting_users_only():
    ws = _converted_ws([
        _row(PurchaseOrderID=1, UpdateUser="jhantz-2"),   # accounting
        _row(PurchaseOrderID=2, UpdateUser="mcohen"),     # look-alike, NOT accounting
        _row(PurchaseOrderID=3, UpdateUser="MaCohen"),    # accounting (case-insensitive)
        _row(PurchaseOrderID=4, UpdateUser="randomuser"), # not accounting
    ])
    hdr = [c.value for c in ws[1]]
    ui = hdr.index("User") + 1
    fills = {ws.cell(r, ui).value: _fill_rgb(ws.cell(r, ui)) for r in range(2, ws.max_row + 1)}
    yellow = processor.ACCOUNTING_FILL.fgColor.rgb
    assert fills["jhantz-2"] == yellow
    assert fills["MaCohen"] == yellow
    assert fills["mcohen"] is None      # must not match macohen
    assert fills["randomuser"] is None


def test_zone1_converted_has_header_filter():
    ws = _converted_ws([_row(PurchaseOrderID=1)])
    assert ws.auto_filter.ref is not None
    assert ws.auto_filter.ref.startswith("A1")


def test_zone1_cancelled_is_yes_blank_not_boolean():
    ws = _converted_ws([
        _row(PurchaseOrderID=1, IsCancelled=True),
        _row(PurchaseOrderID=2, IsCancelled=False),
    ])
    hdr = [c.value for c in ws[1]]
    ci = hdr.index("Cancelled") + 1
    assert ws.cell(2, ci).value == "Yes"
    assert ws.cell(3, ci).value in (None, "")   # blank, not False


def test_zone2_empty_sheet_still_has_header_filter():
    # Every data sheet in the combined workbook carries a header-row filter,
    # including empty per-company tabs.
    import openpyxl
    rows = [_row(PurchaseOrderID=1, CompanyName="YSA",
                 InitialTicketCostTotal=0, TicketCostTotal=100)]
    out = processor.process_files([(_to_xlsx_bytes(rows), "PO_Cost_Changes_2026-05-30.xlsx")])
    res = processor.build_filtered_outputs(
        out["_cleaned"], out["_source_view"], out["_excluded_view"],
        out["date_range"], out["all_companies"],
    )
    wb = openpyxl.load_workbook(io.BytesIO(res["combined"]))
    for name in wb.sheetnames:
        if name == "Summary":
            continue   # grouped pivot layout — intentionally unfiltered
        assert wb[name].auto_filter.ref is not None, f"{name} missing filter"


def test_accounting_users_match_ui_list():
    # Guard against drift between the code constant and the UI reminder list.
    import re
    here = os.path.dirname(__file__)
    html = open(os.path.join(here, "..", "index.html")).read()
    block = re.search(r'upload-warning-users"?>(.*?)</div>', html, re.S).group(1)
    ui_names = {re.sub(r"\(.*?\)", "", s).strip().lower()
                for s in re.findall(r"<span>(.*?)</span>", block)}
    assert ui_names == set(processor.ACCOUNTING_USERS)


def test_zone2_highlights_accounting_users_in_source_data():
    # The User column on Zone 2 Source Data / Excluded tabs highlights
    # accounting-team users, same as the Zone 1 Converted sheet.
    import openpyxl
    rows = [
        _row(PurchaseOrderID=1, CompanyName="YSA", UpdateUser="bblumenthal",
             InitialTicketCostTotal=0, TicketCostTotal=100),   # accounting
        _row(PurchaseOrderID=2, CompanyName="YSA", UpdateUser="someoneelse",
             InitialTicketCostTotal=0, TicketCostTotal=50),    # not accounting
    ]
    out = processor.process_files([(_to_xlsx_bytes(rows), "PO_Cost_Changes_2026-05-30.xlsx")])
    res = processor.build_filtered_outputs(
        out["_cleaned"], out["_source_view"], out["_excluded_view"],
        out["date_range"], out["all_companies"],
    )
    ws = openpyxl.load_workbook(io.BytesIO(res["combined"]))["Source Data"]
    hdr = [c.value for c in ws[1]]
    ui = hdr.index("User") + 1
    yellow = processor.ACCOUNTING_FILL.fgColor.rgb
    fills = {ws.cell(r, ui).value: _fill_rgb(ws.cell(r, ui)) for r in range(2, ws.max_row + 1)}
    assert fills["bblumenthal"] == yellow
    assert fills["someoneelse"] != yellow


# ---------------------------------------------------------------------------
# Summary sort order + download filename / zip-folder tweaks.
# ---------------------------------------------------------------------------

def test_summary_sorted_by_date_then_vendor():
    import openpyxl, datetime as dt
    ledger = pd.DataFrame([
        {"Company": "YSA", "Vendor": "Zzz", "Description": "d1",
         "Date": dt.datetime(2026, 6, 1), "Total": -10.0},
        {"Company": "YSA", "Vendor": "Aaa", "Description": "d2",
         "Date": dt.datetime(2026, 6, 5), "Total": -20.0},
        {"Company": "YSA", "Vendor": "Bbb", "Description": "d3",
         "Date": dt.datetime(2026, 6, 1), "Total": -30.0},
    ])
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    processor._write_summary_sheet(wb, ledger)
    ws = wb["Summary"]
    seen, r = [], 3
    while ws.cell(r, 1).value != "YSA Total":
        seen.append((ws.cell(r, 4).value, ws.cell(r, 2).value))
        r += 1
    assert seen == [
        (dt.datetime(2026, 6, 1), "Bbb"),   # earliest date, vendor B before Z
        (dt.datetime(2026, 6, 1), "Zzz"),
        (dt.datetime(2026, 6, 5), "Aaa"),   # later date last
    ]


def test_combined_download_name_is_plain():
    import app as appmod
    meta = {
        "date_range": "June 9th 2026",
        "selected_companies": ["GK", "YSA", "TL"],
        "companies": ["GK", "YSA"],
        "bills_companies": ["TL"],
    }
    # Always plain "Combined" — no company names in the filename.
    assert appmod._combined_download_name(meta) == "PO Cost Changes - Combined - June 9th 2026.xlsx"
    empty = {"date_range": "June 9th 2026", "selected_companies": [],
             "companies": [], "bills_companies": []}
    assert appmod._combined_download_name(empty) == "PO Cost Changes - Combined - June 9th 2026.xlsx"


# ---------------------------------------------------------------------------
# Zone 1 Converted: date fields render date-only (no timestamp) while the
# underlying datetime is preserved for the Zone 2 round-trip.
# ---------------------------------------------------------------------------

def test_zone1_date_columns_are_date_only_display():
    import openpyxl, datetime as dt
    rows = [_row(PurchaseOrderID=1,
                 AdjustedDateTimeUTC="2026-06-01T04:51:00",
                 CreatedDate="2026-05-15T22:10:00",
                 EventDate="2026-07-04T19:30:00")]
    b = processor.convert_to_modified(_to_xlsx_bytes(rows), "PO_Cost_Changes.xlsx")
    ws = openpyxl.load_workbook(io.BytesIO(b))["Converted"]
    hdr = [c.value for c in ws[1]]
    for name in ("Adjustment Date", "Event Date", "CreatedDate"):
        cell = ws.cell(2, hdr.index(name) + 1)
        assert cell.number_format == "mm/dd/yyyy", name        # displays date-only
        assert isinstance(cell.value, dt.datetime), name        # real datetime kept
    # Time component is preserved underneath (needed by Zone 2).
    adj = ws.cell(2, hdr.index("Adjustment Date") + 1).value
    assert (adj.hour, adj.minute) == (4, 51)


def test_zone1_date_only_preserves_same_day_exclusion_roundtrip():
    # A row whose PO was created the same US-Central day as the adjustment is
    # excluded. The Converted file must reproduce that exclusion exactly, which
    # only holds if the underlying timestamp survives the date-only formatting.
    rows = [_row(PurchaseOrderID=1, CompanyName="YSA",
                 InitialTicketCostTotal=0, TicketCostTotal=100,
                 AdjustedDateTimeUTC="2026-06-10T15:00:00",
                 CreatedDate="2026-06-10T09:00:00", Remove="")]
    raw_bytes = _to_xlsx_bytes(rows, extra_cols=["Remove"])
    conv = processor.convert_to_modified(raw_bytes, "PO_Cost_Changes_2026-06-10.xlsx")
    raw = processor.process_files([(raw_bytes, "PO_Cost_Changes_2026-06-10.xlsx")])
    mod = processor.process_files([(conv, "x (converted).xlsx")])
    assert raw["excluded"]["row_count"] == 1            # excluded same-day in raw
    assert mod["excluded"] == raw["excluded"]           # ...and identically via Converted


def test_ysa_label_is_asher_value_preserved():
    # YSA's tab/file/checkbox label is "Asher"; the Company-column value,
    # the memo text, and the Summary header all stay "YSA".
    assert processor.file_label("YSA") == "Asher"
    assert "Asher" in processor.DISPLAY_ORDER and "YSA" not in processor.DISPLAY_ORDER

    rows = [_row(PurchaseOrderID=1, CompanyName="YSA", PerformerName="Hamilton",
                 AccountEmail="a@b.com", InitialTicketCostTotal=0, TicketCostTotal=100)]
    res = processor.process_files([(_to_xlsx_bytes(rows), "f.xlsx")])
    # Checkbox list (all_companies) shows the label.
    assert "Asher" in res["all_companies"] and "YSA" not in res["all_companies"]

    out = processor.build_filtered_outputs(
        res["_cleaned"], res["_source_view"], res["_excluded_view"],
        res["date_range"], res["all_companies"],
    )
    # File keys (download names) use the label.
    assert set(out["bills_files"].keys()) == {"Asher"}

    import openpyxl, io as _io
    wb = openpyxl.load_workbook(_io.BytesIO(out["combined"]))
    # Combined per-company tab uses the label...
    assert "Asher" in wb.sheetnames and "YSA" not in wb.sheetnames
    # ...but the data value / memo stays "YSA".
    sd = wb["Source Data"]; ci = [c.value for c in sd[1]].index("Company")
    assert {r[ci] for r in sd.iter_rows(min_row=2, values_only=True) if r[ci]} == {"YSA"}
    bills = wb["Bills"]; mi = [c.value for c in bills[1]].index("Memo")
    assert "Cost Changes (YSA)" in bills.cell(2, mi + 1).value


def test_expense_numbers_are_random_8_digit_like_bill_numbers():
    # Expense # is a unique random 8-digit integer per event (same style as the
    # Bills file's "Bill No."), an expense pair's two legs share one number, and
    # the deterministic grouped/ordered layout is preserved.
    import openpyxl, io as _io
    from collections import Counter
    rows = ([_row(PurchaseOrderID=i, CompanyName="YSA", PerformerName=f"E{i}",
                  AccountEmail=f"e{i}@x.com", InitialTicketCostTotal=100, TicketCostTotal=20)
             for i in range(1, 5)] +                                   # 4 expenses
            [_row(PurchaseOrderID=i, CompanyName="YSA", PerformerName=f"B{i}",
                  AccountEmail=f"e{i}@x.com", InitialTicketCostTotal=0, TicketCostTotal=50)
             for i in range(5, 8)])                                    # 3 bills
    res = processor.process_files([(_to_xlsx_bytes(rows), "f.xlsx")])
    out = processor.build_filtered_outputs(res["_cleaned"], res["_source_view"],
            res["_excluded_view"], res["date_range"], res["all_companies"])
    wb = openpyxl.load_workbook(_io.BytesIO(out["combined"]))

    def expense_nums(sheet):
        hdr = [c.value for c in sheet[1]]; ei = hdr.index("Expense #")
        return [sheet.cell(r, ei + 1).value for r in range(2, sheet.max_row + 1)]

    is8 = lambda n: isinstance(n, int) and 10_000_000 <= n <= 99_999_999

    # Combined tab: one row per event (bills + each expense's IA leg).
    comb = expense_nums(wb["Combined"])
    assert comb and all(is8(n) for n in comb)
    assert len(comb) == len(set(comb))          # globally unique
    assert comb == sorted(comb)                 # ordering preserved

    # Per-company expenses tab: two legs per expense share one number.
    exp = expense_nums(wb["Asher"])
    assert exp and all(is8(n) for n in exp)
    assert all(v == 2 for v in Counter(exp).values())
    assert exp == sorted(exp)


def test_mixed_precision_timestamp_still_same_day_excluded():
    # Regression (Grossman PO 2426257): a row whose AdjustedDateTimeUTC lacks
    # fractional seconds — while other rows in the column have them — must still
    # parse and be caught by the same-day exclusion, not slip through as an
    # unparseable NaT date. Both rows below share the same US-Central created and
    # adjusted day, so both must be excluded.
    rows = [
        _row(PurchaseOrderID=999, CompanyName="YSA", PerformerName="P1", AccountEmail="a@x.com",
             InitialTicketCostTotal=0, TicketCostTotal=50,
             AdjustedDateTimeUTC="2026-06-22T20:05:51.497",   # with milliseconds
             CreatedDate="2026-06-22T18:55:54.487", Remove=""),
        _row(PurchaseOrderID=999, CompanyName="YSA", PerformerName="P2", AccountEmail="b@x.com",
             InitialTicketCostTotal=0, TicketCostTotal=55,
             AdjustedDateTimeUTC="2026-06-22T20:05:51",       # NO milliseconds — the bug case
             CreatedDate="2026-06-22T18:55:54.487", Remove=""),
    ]
    res = processor.process_files([(_to_xlsx_bytes(rows, extra_cols=["Remove"]), "PO_Cost_Changes.xlsx")])
    assert res["excluded"]["row_count"] == 2          # both excluded, incl. the no-ms row
    out = processor.build_filtered_outputs(
        res["_cleaned"], res["_source_view"], res["_excluded_view"],
        res["date_range"], res["all_companies"],
    )
    assert "Asher" not in out["bills_files"]           # nothing survives -> no output files
    assert "Asher" not in out["companies"]


def test_excluded_qbo_company_dropped_even_if_in_master(tmp_path):
    # A QBO company in mapping.EXCLUDED_QBO_COMPANIES must be dropped at load
    # time even when the master file still lists it — so a removed company stays
    # gone regardless of the deployed master (matching rows -> "ignored").
    import mapping
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active
    ws.append(["QBO Company", "TicketVault Company /  Applied Payments Category"])
    ws.append(["Damona & Crew", "Damon and Crew"])   # excluded company
    ws.append(["YS Needle Tickets", "Needle YS"])    # excluded company
    ws.append(["Y&S Tickets", "Y&S"])                # normal company, must survive
    path = tmp_path / "master.xlsx"; wb.save(path)
    m = mapping.load_mapping(path)
    assert "damon and crew" not in m
    assert "needle ys" not in m
    assert "Damona & Crew" not in set(m.values())
    assert "YS Needle Tickets" not in set(m.values())
    assert m.get("y&s") == "Y&S Tickets"


def test_season_tag_values():
    # Vendor-based auto-tag, ported from the Purchase Details app.
    assert processor._season_tag("Live Nation Extras") == "LN Extras"
    assert processor._season_tag("Live Nation The Fillmore Detroit") == "Live Nation"
    assert processor._season_tag("Broadway Extras") == "Broadway Extras"
    assert processor._season_tag("Broadway Across America") == "Broadway"
    assert processor._season_tag("Ticketmaster") == ""
    assert processor._season_tag("Boston Red Sox") == ""


def test_seasons_autotagged_on_bills_and_inventory_asset_expense_lines():
    rows = [
        # Concert Extras -> "Live Nation Extras"; positive adjustment -> bill
        _row(PurchaseOrderID=1, CompanyName="YSA", Vendor="Concert Extras",
             VenueName="Some Arena", PerformerName="Act", AccountEmail="a@b.com",
             InitialTicketCostTotal=0, TicketCostTotal=100),
        # Concert Extras; negative adjustment -> expense pair (Line A + offset)
        _row(PurchaseOrderID=2, CompanyName="YSA", Vendor="Concert Extras",
             VenueName="Some Arena", PerformerName="Act2", AccountEmail="c@d.com",
             InitialTicketCostTotal=100, TicketCostTotal=20),
    ]
    cleaned, _ = processor.transform(_norm(rows))
    assert (cleaned["Vendor"] == "Live Nation Extras").all()   # sanity

    bills, expenses = processor._build_bills_and_expenses(cleaned)
    assert "Seasons" in bills.columns and "Seasons" in expenses.columns
    assert (bills["Seasons"] == "LN Extras").all()             # bill tagged
    ia = expenses[expenses["Category"] == "Inventory Asset"]
    off = expenses[expenses["Category"] != "Inventory Asset"]
    assert (ia["Seasons"] == "LN Extras").all()                # inventory-asset leg tagged
    assert (off["Seasons"] == "").all()                        # offset leg blank

    pdb = processor.build_pd_bills(cleaned)                     # PD bills tagged too
    assert (pdb["Seasons"] == "LN Extras").all()


def test_league_for_team_lookup():
    import teams
    teams.reset_cache()
    assert teams.league_for_team("Los Angeles Lakers") == "NBA"
    assert teams.league_for_team("Houston Rockets") == "NBA"
    assert teams.league_for_team("Nobody In Particular") == ""
    assert teams.league_for_team("State University Basketball") == "College"


def _season_rows(event_dates, **over):
    common = dict(CompanyName="Jacks YS", Vendor="Los Angeles Lakers",
                  PerformerName="Los Angeles Lakers", Section="205", Row="10",
                  StartSeat=5, EndSeat=6, AccountEmail="x@y.com",
                  InitialTicketCostTotal=300, TicketCostTotal=0,  # negative -> expense
                  CreatedDate=None)
    common.update(over)
    return [_row(PurchaseOrderID=2308881, EventDate=d, **common) for d in event_dates]


def test_season_ticket_league_tagged_on_inventory_asset_line():
    # Same seat across 3 distinct event dates -> season-ticket group -> NBA on
    # the inventory-asset expense leg; the offset leg stays blank.
    rows = _season_rows(["2026-05-20", "2026-06-05", "2026-06-13"])
    res = processor.process_files([(_to_xlsx_bytes(rows), "PO_Cost_Changes_2026-06-29.xlsx")])
    _, expenses = processor._build_bills_and_expenses(res["_cleaned"])
    ia = expenses[expenses["Category"] == "Inventory Asset"]
    off = expenses[expenses["Category"] != "Inventory Asset"]
    assert len(ia) >= 1 and (ia["Seasons"] == "NBA").all()
    assert (off["Seasons"] == "").all()


def test_two_event_dates_is_not_a_season_ticket():
    # Only 2 distinct event dates -> below the 3-date threshold -> no league tag.
    rows = _season_rows(["2026-05-20", "2026-06-05"])
    res = processor.process_files([(_to_xlsx_bytes(rows), "PO_Cost_Changes_2026-06-29.xlsx")])
    _, expenses = processor._build_bills_and_expenses(res["_cleaned"])
    ia = expenses[expenses["Category"] == "Inventory Asset"]
    assert (ia["Seasons"] == "").all()


def test_season_ticket_excluded_vendor_not_tagged():
    # A resale-marketplace vendor is never league-tagged even if it spans 3+ dates.
    rows = _season_rows(["2026-05-20", "2026-06-05", "2026-06-13"], Vendor="StubHub")
    res = processor.process_files([(_to_xlsx_bytes(rows), "PO_Cost_Changes_2026-06-29.xlsx")])
    _, expenses = processor._build_bills_and_expenses(res["_cleaned"])
    ia = expenses[expenses["Category"] == "Inventory Asset"]
    assert (ia["Seasons"] == "").all()
